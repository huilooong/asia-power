#!/usr/bin/env python3
"""Import 汽修宝 (QXB) bulk vehicle export (.xlsx) into the model-dictionary.

Source columns: 品牌名称 | 车型(full trim string) | 车辆说明 | 车辆图片(comma URLs)
Enriches data/knowledge-base/model-dictionary.json (keyed brandSlug -> 中文车型 -> record).
Additive + source-tagged ("qxb_excel_import"); never deletes. Default = dry-run preview.

Usage:
  python scripts/import-qxb-vehicles.py <file.xlsx>            # dry-run -> reports/qxb-model-import-preview.json
  python scripts/import-qxb-vehicles.py <file.xlsx> --apply    # merge into model-dictionary.json
"""
import json, re, sys, collections, datetime, pathlib
from openpyxl import load_workbook

ROOT = pathlib.Path(__file__).resolve().parents[1]
KB = ROOT / "data" / "knowledge-base"
MODEL_DICT = KB / "model-dictionary.json"
BRAND_DICT = KB / "brand-dictionary.json"

# manufacturer/JV prefixes that sit between brand and the real model name
MFR_PREFIX = ["北京现代", "东风", "一汽", "广汽", "上汽", "长安", "华晨", "进口"]
YEAR = re.compile(r"^\d{4}款?$")
DISP = re.compile(r"^\d\.\d[TL]?$|^\d\.\d$")
STOP = set("MT AT CVT DCT AMT DSG 手动 自动 三厢 两厢 旅行版 旅行 掀背 双擎 混动".split())
GRADE = re.compile(r"^[A-Z]{1,3}L?级$")  # Benz E级/C级/GLA级 etc -> keep whole

def now():
    return datetime.datetime.now(datetime.timezone.utc).isoformat()

def parse_model(brand, trim):
    """Return (model_key_cn_or_latin, english_suggestion, variant, raw)."""
    variant = None
    s = trim
    # capture & strip 进口 marker (parenthesised or bare) -> variant flag
    if re.search(r"[((]进口[))]", s):
        variant = "进口"; s = re.sub(r"[((]进口[))]", "", s)
    toks = s.split()
    out, started = [], False
    for t in toks:
        if not started:
            if t == brand:
                continue
            if t.startswith(brand) and t != brand:
                t = t[len(brand):]
            # strip manufacturer/JV prefix glued to model (北京现代ix35)
            for p in MFR_PREFIX:
                if t.startswith(p) and t != p:
                    t = t[len(p):]
            if t in MFR_PREFIX or not t:
                continue
            started = True
        if YEAR.search(t) or "款" in t:
            break
        if DISP.match(t) or t in STOP:
            break
        out.append(t)
    raw = "".join(out).strip()
    raw = re.sub(r"[((]进口[))]", "", raw).strip()
    if not raw:
        return None, None, variant, "".join(out)
    # Benz grade (E级/C级): keep whole as model, no split
    if GRADE.match(raw):
        return raw, None, variant, raw
    has_cjk = bool(re.search(r"[一-鿿]", raw))
    has_lat = bool(re.search(r"[A-Za-z]", raw))
    if has_cjk and has_lat:                       # 骐达TIIDA -> 骐达 / TIIDA
        cjk = "".join(re.findall(r"[一-鿿]+", raw))
        lat = " ".join(re.findall(r"[A-Za-z0-9\-]+", raw))
        return cjk, lat, variant, raw
    if has_lat and not has_cjk:                   # CR-V / K2 -> latin model
        return raw, raw, variant, raw
    return raw, None, variant, raw                # pure CJK

def main():
    if len(sys.argv) < 2:
        print(__doc__); sys.exit(1)
    src = sys.argv[1]
    apply = "--apply" in sys.argv
    bd = json.load(open(BRAND_DICT))
    md = json.load(open(MODEL_DICT))
    cn2slug = {cn: (v.get("english") or "").lower().replace(" & ", "-").replace(" ", "-")
               for cn, v in bd.items()}
    wb = load_workbook(src, read_only=True, data_only=True); ws = wb.active
    agg = {}
    rows = 0
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        rows += 1
        brand = str(r[0]).strip(); trim = str(r[1] or "").strip()
        slug = cn2slug.get(brand)
        if not slug:
            continue
        key, en, variant, raw = parse_model(brand, trim)
        if not key:
            continue
        yrs = set(re.findall(r"(\d{4})款", trim))
        a = agg.setdefault((slug, key), {"years": set(), "english": None, "variants": set(),
                                         "count": 0, "raw": raw, "sample": trim})
        a["years"] |= yrs; a["count"] += 1
        if en and not a["english"]:
            a["english"] = en
        if variant:
            a["variants"].add(variant)

    new, upd = [], []
    for (slug, key), a in sorted(agg.items()):
        existing = key in md.get(slug, {})
        rec = {"brandSlug": slug, "model": key, "english": a["english"],
               "years": sorted(a["years"]), "variants": sorted(a["variants"]),
               "count": a["count"], "sample": a["sample"]}
        (upd if existing else new).append(rec)

    if not apply:
        out = ROOT / "reports" / "qxb-model-import-preview.json"
        out.parent.mkdir(exist_ok=True)
        json.dump({"source": src, "rows": rows, "new": new, "update": upd},
                  open(out, "w"), ensure_ascii=False, indent=2)
        print(f"DRY-RUN rows={rows} NEW={len(new)} UPDATE={len(upd)} -> {out}")
        for r in new:
            v = ("+" + ",".join(r["variants"])) if r["variants"] else ""
            print(f"  NEW {r['brandSlug']}/{r['model']}{v}  en={r['english'] or '-'}  yrs={','.join(r['years']) or '-'}  n={r['count']}")
        return

    # APPLY: additive merge, source-tagged
    for (slug, key), a in agg.items():
        md.setdefault(slug, {})
        cur = md[slug].get(key, {})
        rec = dict(cur)
        rec["chinese"] = rec.get("chinese") or key
        if a["english"] and not rec.get("english"):
            rec["english"] = a["english"]
        elif "english" not in rec:
            rec["english"] = None
        # merge years
        yrs = set(rec.get("years") or []) | a["years"]
        if yrs:
            rec["years"] = sorted(yrs)
        if a["variants"]:
            rec["variants"] = sorted(set(rec.get("variants") or []) | a["variants"])
        if not cur:
            rec["source"] = "qxb_excel_import"
        rec["learnedAt"] = now()
        md[slug][key] = rec
    tmp = MODEL_DICT.with_suffix(".json.tmp")
    json.dump(md, open(tmp, "w"), ensure_ascii=False, indent=2)
    tmp.replace(MODEL_DICT)
    print(f"APPLIED rows={rows} NEW={len(new)} UPDATE={len(upd)} -> {MODEL_DICT}")

if __name__ == "__main__":
    main()
