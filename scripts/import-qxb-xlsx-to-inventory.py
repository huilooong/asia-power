#!/usr/bin/env python3
"""Import a 汽修宝 Excel export into QXB inventory pipeline (photos + VIN + upload).

New exports often reuse row numbers 2-101 for different vehicles. By default we
append at the next free row (after existing Excel rows) to avoid clashing with
already-uploaded QXB stock IDs.

Usage:
  python scripts/import-qxb-xlsx-to-inventory.py /path/to/file.xlsx --prepare-only
  python scripts/import-qxb-xlsx-to-inventory.py /path/to/file.xlsx --download-only
  python scripts/import-qxb-xlsx-to-inventory.py /path/to/file.xlsx --upload --limit 5
"""

from __future__ import annotations

import argparse
import csv
import json
import shutil
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

from typing import Any

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from dotenv import load_dotenv

load_dotenv(ROOT / ".env")

from inventory_core import qxb_pipeline

ACTIVE_XLSX = ROOT / "data/qxb-vehicles-active.xlsx"
VIN_CSV = ROOT / "reports/qxb-vin-ocr-results.csv"
HEADER = ("品牌名称", "车型", "车辆说明", "车辆图片", "车架号")


def _read_source_rows(path: Path) -> list[tuple]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows: list[tuple] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        brand = str(row[0]).strip()
        trim = str(row[1] or "").strip()
        if not brand or not trim:
            continue
        desc = str(row[2] or "").strip()
        photos = str(row[3] or "").strip()
        vin = str(row[4] or "").strip() if len(row) > 4 else ""
        rows.append((brand, trim, desc, photos, vin))
    wb.close()
    return rows


def _next_start_row() -> int:
    old = Path("/Users/longhui/Downloads/汽修宝车辆数据_2606291619.xlsx")
    if not old.is_file():
        old = ROOT / "data/qxb-vehicles.xlsx"
    existing = qxb_pipeline.load_rows(old) if old.is_file() else []
    if existing:
        return max(r["row"] for r in existing) + 1
    return 2


def prepare_active_xlsx(source: Path, start_row: int | None) -> dict:
    items = _read_source_rows(source)
    if not items:
        raise SystemExit(f"No vehicle rows found in {source}")

    begin = start_row or _next_start_row()

    wb = Workbook()
    ws = wb.active
    ws.title = "SheetJS"
    ws.append(HEADER)
    for offset, item in enumerate(items):
        excel_row = begin + offset
        for col, value in enumerate(item, start=1):
            ws.cell(row=excel_row, column=col, value=value)
    ACTIVE_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ACTIVE_XLSX)
    wb.close()

    row_ids = list(range(begin, begin + len(items)))
    return {
        "source": str(source),
        "active_xlsx": str(ACTIVE_XLSX),
        "start_row": begin,
        "end_row": begin + len(items) - 1,
        "count": len(items),
        "row_ids": row_ids,
        "stock_ids": [f"QXB{r:04d}" for r in row_ids],
    }


def import_vins(meta: dict) -> int:
    items = _read_source_rows(Path(meta["source"]))
    row_ids = meta["row_ids"]
    if len(items) != len(row_ids):
        raise RuntimeError("source row count mismatch after prepare")

    existing: dict[int, dict[str, str]] = {}
    if VIN_CSV.is_file():
        for rec in qxb_pipeline._read_vin_csv_rows(VIN_CSV):
            existing[int(rec["row"])] = rec

    added = 0
    for row_id, item in zip(row_ids, items):
        vin = qxb_pipeline.normalize_vin_strict(item[4])
        if not vin or not qxb_pipeline.VIN_STRICT.match(vin):
            continue
        trim = item[1]
        existing[row_id] = {
            "row": str(row_id),
            "model": trim,
            "vin": vin,
            "image_path": "excel_vin_column",
            "confidence": "excel",
            "engine_code": "",
            "transmission_code": "",
        }
        added += 1

    qxb_pipeline._write_vin_csv_rows(
        VIN_CSV,
        [existing[k] for k in sorted(existing)],
    )
    return added


def validate_import_collisions(meta: dict) -> list[dict[str, Any]]:
    """Warn when Excel VINs already exist on another QXB row (prevents silent dup imports)."""
    ctx = qxb_pipeline.load_context()
    warnings: list[dict[str, Any]] = []
    items = _read_source_rows(Path(meta["source"]))
    for row_id, item in zip(meta["row_ids"], items):
        vin = qxb_pipeline.normalize_vin_strict(item[4])
        if not vin:
            continue
        dup = qxb_pipeline.find_duplicate_vin(ctx, vin, exclude_row=row_id)
        if dup:
            warnings.append(
                {
                    "row": row_id,
                    "stockId": f"QXB{row_id:04d}",
                    "vin": vin,
                    "duplicateOf": dup,
                }
            )
    return warnings


def run_download(source_for_photos: Path) -> None:
    cmd = [
        sys.executable,
        str(ROOT / "scripts/download-qxb-photos.py"),
        str(source_for_photos),
        "--workers",
        "8",
        "--timeout",
        "25",
        "--retries",
        "2",
    ]
    subprocess.run(cmd, cwd=str(ROOT), check=True)


def run_upload_rows(row_ids: list[int]) -> list[dict]:
    from importlib.machinery import SourceFileLoader

    batch = SourceFileLoader(
        "qxb_batch_upload",
        str(ROOT / "scripts/qxb-batch-upload.py"),
    ).load_module()
    return [batch.upload_row(row) for row in row_ids]


def main() -> int:
    parser = argparse.ArgumentParser(description="Import QXB Excel into inventory pipeline")
    parser.add_argument("xlsx", type=Path)
    parser.add_argument("--start-row", type=int, default=0, help="Force first Excel row id")
    parser.add_argument("--prepare-only", action="store_true")
    parser.add_argument("--download-only", action="store_true")
    parser.add_argument("--skip-download", action="store_true")
    parser.add_argument("--skip-vins", action="store_true")
    parser.add_argument("--upload", action="store_true")
    parser.add_argument("--limit", type=int, default=0, help="Max rows to upload this run")
    args = parser.parse_args()

    if not args.xlsx.is_file():
        print(f"Missing file: {args.xlsx}")
        return 1

    meta = prepare_active_xlsx(args.xlsx, args.start_row or None)
    (ROOT / "reports/qxb-active-import.json").write_text(
        json.dumps(meta, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(json.dumps({"stage": "prepared", **meta}, ensure_ascii=False, indent=2))

    if args.prepare_only:
        return 0

    if not args.skip_vins:
        vin_count = import_vins(meta)
        print(json.dumps({"stage": "vins", "imported": vin_count}, ensure_ascii=False))
        collisions = validate_import_collisions(meta)
        if collisions:
            warn_path = ROOT / "reports/qxb-active-import-collisions.json"
            warn_path.write_text(
                json.dumps(collisions, ensure_ascii=False, indent=2) + "\n",
                encoding="utf-8",
            )
            print(
                json.dumps(
                    {
                        "stage": "collision_warning",
                        "count": len(collisions),
                        "report": str(warn_path),
                        "message": "部分 VIN 已在其他 QXB 行上传过，请确认行号后再批量上传",
                    },
                    ensure_ascii=False,
                )
            )

    if not args.skip_download:
        run_download(ACTIVE_XLSX)

    if args.download_only:
        return 0

    if args.upload:
        rows = meta["row_ids"]
        if args.limit > 0:
            rows = rows[: args.limit]
        results = run_upload_rows(rows)
        ok = sum(1 for r in results if r.get("ok"))
        out = ROOT / "reports/qxb-active-import-upload.json"
        out.write_text(json.dumps(results, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        print(json.dumps({"stage": "upload", "ok": ok, "failed": len(results) - ok, "report": str(out)}, ensure_ascii=False))
        return 0 if ok == len(results) else 2

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
