#!/usr/bin/env python3
"""Merge all AsiaPower customer sources into one master table for CEO export."""

from __future__ import annotations

import csv
import json
import re
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

WORKSPACE = Path("/Users/longhui/Desktop/AsiaPower")
DOWNLOADS = Path("/Users/longhui/Downloads")
EXPORT_DATE = "2026-07-04"

MAPS_AFRICA = WORKSPACE / "memory/customer_gateway/africa_maps_leads.jsonl"
MAPS_EARLY = WORKSPACE / "memory/customer_gateway/maps_leads.jsonl"
OUTREACH_DIR = WORKSPACE / "memory/customer_gateway/outreach_queue"
DRAFT_DIR = WORKSPACE / "memory/customer_gateway/draft_queue"
CUSTOMERS_DIR = WORKSPACE / "memory/customers"
CONTACT_LEADS = Path("/tmp/contact-leads-2026-07-04.json")
REPORT_SRC = WORKSPACE / "reports/customer-data-export-2026-07-04.md"

MASTER_COLUMNS = [
    "序号",
    "来源类型",
    "店名/姓名",
    "国家",
    "城市",
    "邮箱",
    "电话",
    "WhatsApp",
    "地址",
    "网站",
    "Google评分",
    "评论数",
    "业务类型/备注",
    "开发信主题",
    "邮件状态",
    "关联邮件数",
    "抓取/提交时间",
    "内部ID",
]

SOURCE_COLUMNS = {
    "maps": [
        "序号",
        "店名",
        "国家",
        "城市",
        "邮箱",
        "电话",
        "地址",
        "网站",
        "Google评分",
        "评论数",
        "业务类型",
        "搜索词",
        "抓取时间",
        "内部ID",
        "Maps链接",
    ],
    "outreach": [
        "序号",
        "店名",
        "国家",
        "邮箱",
        "电话",
        "网站",
        "开发信主题",
        "邮件状态",
        "创建时间",
        "内部ID",
        "备注",
    ],
    "website": [
        "序号",
        "姓名",
        "国家",
        "邮箱",
        "电话",
        "询盘内容",
        "来源页",
        "回复状态",
        "提交时间",
        "内部ID",
        "IP国家",
    ],
    "whatsapp": [
        "序号",
        "客户名",
        "客户原话",
        "拟回复摘要",
        "状态",
        "创建时间",
        "内部ID",
        "chat_id",
    ],
    "profiles": [
        "序号",
        "客户标识",
        "国家",
        "类型",
        "需求产品",
        "跟进状态",
        "最后联系",
        "来源",
        "文件",
    ],
}


def norm_email(value: str | None) -> str:
    if not value:
        return ""
    return value.strip().lower()


def norm_phone(value: str | None) -> str:
    if not value:
        return ""
    digits = re.sub(r"\D", "", value)
    if digits.startswith("00"):
        digits = digits[2:]
    return digits


def norm_name(value: str | None) -> str:
    if not value:
        return ""
    text = value.strip().lower()
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"[^\w\s\-&'.]", "", text)
    return text


def norm_country(value: str | None) -> str:
    if not value:
        return ""
    return value.strip().lower()


def dedup_key(email: str, name: str, country: str) -> str:
    email_n = norm_email(email)
    if email_n:
        return f"email:{email_n}"
    name_n = norm_name(name)
    country_n = norm_country(country)
    if name_n and country_n:
        return f"name_country:{name_n}|{country_n}"
    if name_n:
        return f"name:{name_n}"
    return ""


def first_non_empty(*values: str | None) -> str:
    for value in values:
        if value and str(value).strip():
            return str(value).strip()
    return ""


def load_jsonl(path: Path) -> list[dict]:
    rows: list[dict] = []
    if not path.exists():
        return rows
    with path.open(encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if not line:
                continue
            rows.append(json.loads(line))
    return rows


def load_maps_africa() -> list[dict]:
    rows = []
    for idx, item in enumerate(load_jsonl(MAPS_AFRICA), 1):
        rows.append(
            {
                "source_type": "Maps",
                "name": item.get("business_name", ""),
                "country": item.get("country", ""),
                "city": item.get("city", ""),
                "email": item.get("email", ""),
                "phone": item.get("phone", ""),
                "whatsapp": "",
                "address": item.get("address", ""),
                "website": item.get("website", ""),
                "rating": item.get("rating") or "",
                "review_count": item.get("review_count") or item.get("reviews") or "",
                "notes": first_non_empty(
                    item.get("category", ""),
                    item.get("source_query", ""),
                ),
                "subject": "",
                "email_status": "",
                "timestamp": item.get("saved_at", ""),
                "internal_id": item.get("lead_key", f"maps-africa-{idx}"),
                "extra": item.get("maps_url", ""),
            }
        )
    return rows


def load_maps_early() -> list[dict]:
    rows = []
    for idx, item in enumerate(load_jsonl(MAPS_EARLY), 1):
        rows.append(
            {
                "source_type": "Maps(早期试点)",
                "name": item.get("name") or item.get("business_name", ""),
                "country": item.get("country", ""),
                "city": item.get("city", ""),
                "email": item.get("email", ""),
                "phone": item.get("phone", ""),
                "whatsapp": "",
                "address": item.get("address", ""),
                "website": item.get("website", ""),
                "rating": item.get("rating") or "",
                "review_count": item.get("review_count") or "",
                "notes": first_non_empty(item.get("query", ""), item.get("source", "")),
                "subject": "",
                "email_status": "",
                "timestamp": item.get("saved_at", ""),
                "internal_id": item.get("lead_key", f"maps-early-{idx}"),
                "extra": item.get("maps_url", ""),
            }
        )
    return rows


def load_outreach() -> list[dict]:
    rows = []
    for path in sorted(OUTREACH_DIR.glob("*.json")):
        item = json.loads(path.read_text(encoding="utf-8"))
        candidate = item.get("candidate") or {}
        phone = candidate.get("phone", "")
        if phone.startswith("tel:"):
            phone = phone[4:]
        rows.append(
            {
                "source_type": "开发信",
                "name": candidate.get("name", ""),
                "country": candidate.get("country", ""),
                "city": "",
                "email": candidate.get("email", ""),
                "phone": phone,
                "whatsapp": "",
                "address": "",
                "website": candidate.get("website", ""),
                "rating": "",
                "review_count": "",
                "notes": candidate.get("reason", ""),
                "subject": item.get("email_subject", ""),
                "email_status": item.get("status", ""),
                "timestamp": item.get("created_at", ""),
                "internal_id": item.get("outreach_id", path.stem),
                "extra": candidate.get("candidate_id", ""),
            }
        )
    return rows


def load_website() -> list[dict]:
    if not CONTACT_LEADS.exists():
        return []
    data = json.loads(CONTACT_LEADS.read_text(encoding="utf-8"))
    leads = data if isinstance(data, list) else data.get("leads", [])
    rows = []
    for item in leads:
        enquiry = first_non_empty(
            item.get("vehicleDetails", ""),
            item.get("message", ""),
            f"{item.get('brand', '')} {item.get('model', '')} {item.get('engineCode', '')}".strip(),
        )
        country = first_non_empty(item.get("ipCountry"), item.get("country"))
        rows.append(
            {
                "source_type": "网站表单",
                "name": first_non_empty(item.get("name"), item.get("company"), "(未留名)"),
                "country": country,
                "city": item.get("ipCity", ""),
                "email": item.get("email", ""),
                "phone": item.get("phone", ""),
                "whatsapp": item.get("phone", "") if item.get("replyChannel") == "whatsapp" else "",
                "address": "",
                "website": item.get("pageUrl", ""),
                "rating": "",
                "review_count": "",
                "notes": enquiry,
                "subject": item.get("inquirySubject", ""),
                "email_status": item.get("replyStatus", ""),
                "timestamp": item.get("createdAt", ""),
                "internal_id": item.get("id", ""),
                "extra": item.get("source", ""),
            }
        )
    return rows


def load_whatsapp_drafts() -> list[dict]:
    rows = []
    for path in sorted(DRAFT_DIR.glob("*.json")):
        item = json.loads(path.read_text(encoding="utf-8"))
        draft = (item.get("customer_reply_draft") or "")[:120].replace("\n", " ")
        rows.append(
            {
                "source_type": "WhatsApp",
                "name": item.get("customer_name", ""),
                "country": "",
                "city": "",
                "email": "",
                "phone": "",
                "whatsapp": item.get("customer_name", ""),
                "address": "",
                "website": "",
                "rating": "",
                "review_count": "",
                "notes": item.get("original_message", ""),
                "subject": "",
                "email_status": item.get("status", ""),
                "timestamp": item.get("created_at", ""),
                "internal_id": item.get("draft_id", path.stem),
                "extra": draft,
                "chat_id": item.get("chat_id", ""),
            }
        )
    return rows


def parse_customer_md(path: Path) -> dict | None:
    text = path.read_text(encoding="utf-8")
    if path.name == ".gitkeep":
        return None
    latest: dict[str, str] = {}
    for line in text.splitlines():
        m = re.match(r"^- \*\*(.+?):\*\* (.+)$", line.strip())
        if m:
            latest[m.group(1).strip()] = m.group(2).strip()
    if not latest:
        return None
    customer = latest.get("Customer", path.stem)
    return {
        "source_type": "客户画像",
        "name": customer,
        "country": latest.get("Country", ""),
        "city": "",
        "email": "",
        "phone": customer if customer.startswith("+") or customer[0:1].isdigit() else "",
        "whatsapp": customer if "+" in customer or re.search(r"\d{3}", customer) else "",
        "address": "",
        "website": "",
        "rating": "",
        "review_count": "",
        "notes": latest.get("Conversation Summary", ""),
        "subject": "",
        "email_status": latest.get("Follow-up Status", ""),
        "timestamp": latest.get("Last Contact", ""),
        "internal_id": path.stem,
        "buyer_type": latest.get("Buyer or Supplier", ""),
        "products": latest.get("Interested Products", ""),
        "profile_source": latest.get("Source", ""),
        "profile_file": path.name,
    }


def load_profiles() -> list[dict]:
    rows = []
    for path in sorted(CUSTOMERS_DIR.glob("*.md")):
        parsed = parse_customer_md(path)
        if parsed:
            rows.append(parsed)
    return rows


def merge_records(all_rows: list[dict]) -> list[dict]:
    buckets: dict[str, list[dict]] = defaultdict(list)
    orphan: list[dict] = []

    for row in all_rows:
        key = dedup_key(row.get("email", ""), row.get("name", ""), row.get("country", ""))
        if key:
            buckets[key].append(row)
        else:
            orphan.append(row)

    merged: list[dict] = []
    for key, group in buckets.items():
        group.sort(key=lambda r: r.get("timestamp") or "")
        base = dict(group[-1])
        source_types = sorted({r.get("source_type", "") for r in group if r.get("source_type")})
        base["source_type"] = "/".join(source_types)
        base["email_count"] = sum(1 for r in group if r.get("source_type") == "开发信")
        if base["email_count"] == 0:
            base["email_count"] = ""

        subjects = [r.get("subject", "") for r in group if r.get("subject")]
        base["subject"] = subjects[-1] if subjects else base.get("subject", "")

        statuses = [r.get("email_status", "") for r in group if r.get("email_status")]
        base["email_status"] = statuses[-1] if statuses else base.get("email_status", "")

        internal_ids = [r.get("internal_id", "") for r in group if r.get("internal_id")]
        base["internal_id"] = internal_ids[0] if len(internal_ids) == 1 else "; ".join(internal_ids[:3]) + (
            f" (+{len(internal_ids)-3})" if len(internal_ids) > 3 else ""
        )

        for field in ("name", "country", "city", "email", "phone", "whatsapp", "address", "website", "rating", "review_count", "notes"):
            base[field] = first_non_empty(*[r.get(field, "") for r in group])

        merged.append(base)

    for row in orphan:
        row = dict(row)
        row["email_count"] = 1 if row.get("source_type") == "开发信" else ""
        merged.append(row)

    merged.sort(key=lambda r: (norm_country(r.get("country", "")), norm_name(r.get("name", ""))))
    return merged


def row_to_master(idx: int, row: dict) -> dict:
    return {
        "序号": idx,
        "来源类型": row.get("source_type", ""),
        "店名/姓名": row.get("name", ""),
        "国家": row.get("country", ""),
        "城市": row.get("city", ""),
        "邮箱": row.get("email", ""),
        "电话": row.get("phone", ""),
        "WhatsApp": row.get("whatsapp", ""),
        "地址": row.get("address", ""),
        "网站": row.get("website", ""),
        "Google评分": row.get("rating", ""),
        "评论数": row.get("review_count", ""),
        "业务类型/备注": row.get("notes", ""),
        "开发信主题": row.get("subject", ""),
        "邮件状态": row.get("email_status", ""),
        "关联邮件数": row.get("email_count", ""),
        "抓取/提交时间": row.get("timestamp", ""),
        "内部ID": row.get("internal_id", ""),
    }


def write_csv(path: Path, columns: list[str], rows: list[dict]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def country_stats(rows: list[dict]) -> list[dict]:
    counter = Counter()
    email_counter = Counter()
    for row in rows:
        country = row.get("country") or "(未知)"
        counter[country] += 1
        if row.get("email"):
            email_counter[country] += 1
    stats = []
    for idx, (country, count) in enumerate(counter.most_common(), 1):
        stats.append(
            {
                "排名": idx,
                "国家": country,
                "客户数": count,
                "有邮箱": email_counter[country],
            }
        )
    return stats


def build_source_tables(
    maps_rows: list[dict],
    outreach_rows: list[dict],
    website_rows: list[dict],
    whatsapp_rows: list[dict],
    profile_rows: list[dict],
) -> dict[str, list[dict]]:
    maps_table = []
    for idx, row in enumerate(maps_rows, 1):
        maps_table.append(
            {
                "序号": idx,
                "店名": row["name"],
                "国家": row["country"],
                "城市": row["city"],
                "邮箱": row["email"],
                "电话": row["phone"],
                "地址": row["address"],
                "网站": row["website"],
                "Google评分": row["rating"],
                "评论数": row["review_count"],
                "业务类型": row["notes"].split(" · ")[0] if row["notes"] else "",
                "搜索词": row["notes"],
                "抓取时间": row["timestamp"],
                "内部ID": row["internal_id"],
                "Maps链接": row.get("extra", ""),
            }
        )

    outreach_table = []
    for idx, row in enumerate(outreach_rows, 1):
        outreach_table.append(
            {
                "序号": idx,
                "店名": row["name"],
                "国家": row["country"],
                "邮箱": row["email"],
                "电话": row["phone"],
                "网站": row["website"],
                "开发信主题": row["subject"],
                "邮件状态": row["email_status"],
                "创建时间": row["timestamp"],
                "内部ID": row["internal_id"],
                "备注": row["notes"],
            }
        )

    website_table = []
    for idx, row in enumerate(website_rows, 1):
        website_table.append(
            {
                "序号": idx,
                "姓名": row["name"],
                "国家": row["country"],
                "邮箱": row["email"],
                "电话": row["phone"],
                "询盘内容": row["notes"],
                "来源页": row.get("extra", ""),
                "回复状态": row["email_status"],
                "提交时间": row["timestamp"],
                "内部ID": row["internal_id"],
                "IP国家": row["city"],
            }
        )

    whatsapp_table = []
    for idx, row in enumerate(whatsapp_rows, 1):
        whatsapp_table.append(
            {
                "序号": idx,
                "客户名": row["name"],
                "客户原话": row["notes"],
                "拟回复摘要": row.get("extra", ""),
                "状态": row["email_status"],
                "创建时间": row["timestamp"],
                "内部ID": row["internal_id"],
                "chat_id": row.get("chat_id", ""),
            }
        )

    profile_table = []
    for idx, row in enumerate(profile_rows, 1):
        profile_table.append(
            {
                "序号": idx,
                "客户标识": row["name"],
                "国家": row["country"],
                "类型": row.get("buyer_type", ""),
                "需求产品": row.get("products", ""),
                "跟进状态": row["email_status"],
                "最后联系": row["timestamp"],
                "来源": row.get("profile_source", ""),
                "文件": row.get("profile_file", row["internal_id"]),
            }
        )

    return {
        "maps": maps_table,
        "outreach": outreach_table,
        "website": website_table,
        "whatsapp": whatsapp_table,
        "profiles": profile_table,
    }


def write_xlsx(path: Path, master_rows: list[dict], source_tables: dict[str, list[dict]], stats_rows: list[dict]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    wb.remove(wb.active)

    sheets = [
        ("全部客户", MASTER_COLUMNS, master_rows),
        ("Maps抓取", SOURCE_COLUMNS["maps"], source_tables["maps"]),
        ("开发邮件", SOURCE_COLUMNS["outreach"], source_tables["outreach"]),
        ("网站线索", SOURCE_COLUMNS["website"], source_tables["website"]),
        ("WhatsApp草稿", SOURCE_COLUMNS["whatsapp"], source_tables["whatsapp"]),
        ("按国家统计", ["排名", "国家", "客户数", "有邮箱"], stats_rows),
    ]

    for title, columns, rows in sheets:
        ws = wb.create_sheet(title)
        ws.append(columns)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in rows:
            ws.append([row.get(col, "") for col in columns])
        ws.freeze_panes = "A2"

    wb.save(path)


def build_report(master_count: int, raw_counts: dict[str, int], merged_email_dupes: int) -> str:
    base = REPORT_SRC.read_text(encoding="utf-8") if REPORT_SRC.exists() else ""
    appendix = f"""

---

## 6. 完整大表说明（2026-07-04 导出）

### 6.1 文件清单（Downloads 文件夹）

| 文件 | 说明 |
|------|------|
| `AsiaPower-客户资料完整表-2026-07-04.csv` | **去重后完整大表**（{master_count} 行） |
| `AsiaPower-客户资料完整表-2026-07-04.xlsx` | 多 Sheet：全部客户 / Maps / 开发邮件 / 网站线索 / WhatsApp / 国家统计 |
| `AsiaPower-网站线索-2026-07-04.csv` | 网站表单 {raw_counts['website']} 条（未去重） |
| `AsiaPower-开发邮件队列-2026-07-04.csv` | 开发邮件 {raw_counts['outreach']} 条（未去重） |
| `AsiaPower-客户资料报告-2026-07-04.md` | 本报告 |

### 6.2 去重逻辑

1. **合并键**：同一 **邮箱**（不区分大小写）**或** 同一 **店名+国家**（标准化后）视为同一客户。
2. **合并方式**：保留最新一条的非空字段；来源类型用 `/` 拼接（如 `Maps/开发信`）。
3. **关联邮件数**：同一客户名下 `开发信` 来源记录数；无开发信时留空。
4. **未去重分表**：网站线索、开发邮件队列 CSV 保留原始条数，便于逐封审批。

### 6.3 原始条数 vs 去重后

| 来源 | 原始条数 | 说明 |
|------|----------|------|
| Maps 非洲抓取 | {raw_counts['maps_africa']} | africa_maps_leads.jsonl |
| Maps 早期试点 | {raw_counts['maps_early']} | maps_leads.jsonl |
| 开发邮件队列 | {raw_counts['outreach']} | outreach_queue/*.json |
| 网站表单 | {raw_counts['website']} | 生产 contact-leads.json |
| WhatsApp 草稿 | {raw_counts['whatsapp']} | draft_queue/*.json |
| 客户画像 | {raw_counts['profiles']} | memory/customers/*.md |
| **去重后合计** | **{master_count}** | 含跨来源合并；{merged_email_dupes} 组客户有 ≥2 封开发信 |

### 6.4 CEO 快速打开

```bash
open ~/Downloads/AsiaPower-客户资料完整表-2026-07-04.csv
open ~/Downloads/AsiaPower-客户资料完整表-2026-07-04.xlsx
```

*完整表生成时间：{EXPORT_DATE} · UTF-8 BOM CSV + openpyxl XLSX*
"""
    return base.rstrip() + appendix


def main() -> None:
    maps_africa = load_maps_africa()
    maps_early = load_maps_early()
    outreach = load_outreach()
    website = load_website()
    whatsapp = load_whatsapp_drafts()
    profiles = load_profiles()

    maps_all = maps_africa + maps_early
    all_raw = maps_all + outreach + website + whatsapp + profiles
    merged = merge_records(all_raw)
    master_rows = [row_to_master(i, row) for i, row in enumerate(merged, 1)]

    source_tables = build_source_tables(maps_all, outreach, website, whatsapp, profiles)
    stats_rows = country_stats(master_rows)

    raw_counts = {
        "maps_africa": len(maps_africa),
        "maps_early": len(maps_early),
        "outreach": len(outreach),
        "website": len(website),
        "whatsapp": len(whatsapp),
        "profiles": len(profiles),
    }
    merged_email_dupes = sum(1 for r in master_rows if isinstance(r.get("关联邮件数"), int) and r["关联邮件数"] >= 2)

    DOWNLOADS.mkdir(parents=True, exist_ok=True)

    csv_master = DOWNLOADS / f"AsiaPower-客户资料完整表-{EXPORT_DATE}.csv"
    xlsx_master = DOWNLOADS / f"AsiaPower-客户资料完整表-{EXPORT_DATE}.xlsx"
    csv_website = DOWNLOADS / f"AsiaPower-网站线索-{EXPORT_DATE}.csv"
    csv_outreach = DOWNLOADS / f"AsiaPower-开发邮件队列-{EXPORT_DATE}.csv"
    md_report = DOWNLOADS / f"AsiaPower-客户资料报告-{EXPORT_DATE}.md"

    write_csv(csv_master, MASTER_COLUMNS, master_rows)
    write_csv(csv_website, SOURCE_COLUMNS["website"], source_tables["website"])
    write_csv(csv_outreach, SOURCE_COLUMNS["outreach"], source_tables["outreach"])
    write_xlsx(xlsx_master, master_rows, source_tables, stats_rows)
    md_report.write_text(
        build_report(len(master_rows), raw_counts, merged_email_dupes),
        encoding="utf-8",
    )

    print(json.dumps(
        {
            "master_csv": str(csv_master),
            "master_rows": len(master_rows),
            "master_size": csv_master.stat().st_size,
            "xlsx": str(xlsx_master),
            "xlsx_size": xlsx_master.stat().st_size,
            "website_csv": str(csv_website),
            "website_rows": len(source_tables["website"]),
            "outreach_csv": str(csv_outreach),
            "outreach_rows": len(source_tables["outreach"]),
            "report_md": str(md_report),
            "raw_counts": raw_counts,
            "merged_email_dupes": merged_email_dupes,
        },
        ensure_ascii=False,
        indent=2,
    ))


if __name__ == "__main__":
    main()
