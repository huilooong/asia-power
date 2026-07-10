#!/usr/bin/env python3
"""Build/upload QXB vehicles into AsiaPower approved half-cut inventory.

Pipeline:
  1. Read QXB Excel rows.
  2. Read downloaded photo manifest.
  3. Read OCR VIN results.
  4. Upload selected local photos to production R2 via /api/half-cuts/upload/presign.
  5. Write approved inventory JSON for server-side merge/promote.

This script is resumable. Uploaded media responses are cached in:
  reports/qxb-site-upload-state.json

It does not print supplier keys.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import mimetypes
import os
import pathlib
import re
import time
import urllib.error
import urllib.parse
import urllib.request
from collections import defaultdict

from openpyxl import load_workbook

import sys

ROOT = pathlib.Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from tools import knowledge_ingest
from inventory_core.qxb_photo_pick import pick_photo_slots
DEFAULT_XLSX = pathlib.Path("/Users/longhui/Downloads/汽修宝车辆数据_2606291619.xlsx")
DEFAULT_MANIFEST = ROOT / "data/qxb-photos/manifest.csv"
DEFAULT_VIN = ROOT / "reports/qxb-vin-ocr-results.csv"
DEFAULT_STATE = ROOT / "reports/qxb-site-upload-state.json"
DEFAULT_OUT = ROOT / "reports/qxb-approved-import.json"
MODEL_DICT = ROOT / "data/knowledge-base/model-dictionary.json"
BRAND_DICT = ROOT / "data/knowledge-base/brand-dictionary.json"
API_BASE = "https://asia-power.com"

MFR_PREFIX = ["北京现代", "东风", "一汽", "广汽", "上汽", "长安", "华晨", "进口"]
YEAR_RE = re.compile(r"(\d{4})款")
DISP_RE = re.compile(r"^\d\.\d[TL]?$|^\d\.\d$")
STOP = set("MT AT CVT DCT AMT DSG 手动 自动 三厢 两厢 旅行版 旅行 掀背 双擎 混动".split())

PHOTO_LABELS = [
    "Front view",
    "Rear view",
    "Engine bay",
    "VIN / chassis plate",
    "Interior",
]


def load_json(path: pathlib.Path, fallback):
    if not path.exists():
        return fallback
    return json.loads(path.read_text())


def save_json(path: pathlib.Path, data):
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n")
    tmp.replace(path)


def slugify(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", str(value or "").lower()).strip("-")


def parse_model(brand: str, trim: str) -> tuple[str, str | None]:
    s = trim or ""
    toks = s.split()
    out = []
    started = False
    for token in toks:
        t = token
        if not started:
            if t == brand:
                continue
            if t.startswith(brand) and t != brand:
                t = t[len(brand):]
            for prefix in MFR_PREFIX:
                if t.startswith(prefix) and t != prefix:
                    t = t[len(prefix):]
            if t in MFR_PREFIX or not t:
                continue
            started = True
        if "款" in t or DISP_RE.match(t) or t in STOP:
            break
        out.append(t)
    raw = "".join(out).strip() or trim or ""
    cjk = "".join(re.findall(r"[\u4e00-\u9fff]+", raw))
    latin = " ".join(re.findall(r"[A-Za-z0-9\-]+", raw))
    if cjk and latin:
        return cjk, latin
    if latin and not cjk:
        return latin, latin
    return raw, None


def load_rows(xlsx: pathlib.Path):
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        brand = str(row[0] or "").strip()
        trim = str(row[1] or "").strip()
        desc = str(row[2] or "").strip()
        if not brand or not trim:
            continue
        year_match = YEAR_RE.search(trim)
        yield {
            "row": row_num,
            "brand_cn": brand,
            "trim": trim,
            "description": desc,
            "year": int(year_match.group(1)) if year_match else None,
        }


def load_brand_maps():
    brand_dict = load_json(BRAND_DICT, {})
    cn_to_en = {}
    cn_to_slug = {}
    for cn, value in brand_dict.items():
        en = value.get("english") or cn
        cn_to_en[cn] = en
        cn_to_slug[cn] = slugify(en)
    return cn_to_en, cn_to_slug


def model_english(brand_slug: str, model_key: str, fallback: str | None):
    md = load_json(MODEL_DICT, {})
    rec = md.get(brand_slug, {}).get(model_key) or {}
    return rec.get("english") or fallback or model_key


def load_manifest(path: pathlib.Path):
    by_row = defaultdict(list)
    for rec in csv.DictReader(path.open()):
        if rec.get("status") not in {"downloaded", "exists"}:
            continue
        local = pathlib.Path(rec["local_path"])
        if not local.exists() or local.stat().st_size <= 0:
            continue
        by_row[int(rec["row"])].append(rec)
    for items in by_row.values():
        items.sort(key=lambda r: int(r["image_index"]))
    return by_row


def load_vins(path: pathlib.Path):
    vins = {}
    if not path.exists():
        return vins
    for rec in csv.DictReader(path.open()):
        vin = str(rec.get("vin") or "").strip().upper()
        if vin:
            vins[int(rec["row"])] = {
                "vin": vin,
                "image_path": rec.get("image_path") or "",
                "confidence": rec.get("confidence") or "",
            }
    return vins


def selected_photos(row: int, photos: list[dict], vin_info: dict | None, max_photos: int):
    picks, _meta = pick_photo_slots(photos, vin_info, row=row, max_photos=max_photos)
    return picks


def request_json(url: str, method: str, headers: dict, body: dict | None = None):
    data = None if body is None else json.dumps(body).encode()
    req = urllib.request.Request(url, data=data, method=method, headers={
        "User-Agent": "AsiaPower-QXB-BatchUploader/1.0",
        **headers,
    })
    if body is not None:
        req.add_header("Content-Type", "application/json")
    with urllib.request.urlopen(req, timeout=30) as res:
        return json.loads(res.read().decode())


def upload_file(path: pathlib.Path, label: str, supplier_key: str, token: str):
    mime = mimetypes.guess_type(str(path))[0] or "image/jpeg"
    size = path.stat().st_size
    presign = request_json(
        f"{API_BASE}/api/half-cuts/upload/presign",
        "POST",
        {"X-Supplier-Key": supplier_key, "X-Upload-Token": token},
        {"kind": "photo", "mimeType": mime, "filename": path.name, "size": size, "label": label},
    )
    req = urllib.request.Request(
        presign["uploadUrl"],
        data=path.read_bytes(),
        method="PUT",
        headers={"Content-Type": presign.get("mimeType") or mime},
    )
    with urllib.request.urlopen(req, timeout=120) as res:
        if res.status >= 300:
            raise RuntimeError(f"R2 upload failed: {res.status}")
    return {
        "label": label,
        "url": presign["url"],
        "fileName": path.name,
        "mimeType": mime,
        "size": size,
    }


def get_upload_token(supplier_key: str):
    return request_json(
        f"{API_BASE}/api/half-cuts/upload-token",
        "POST",
        {"X-Supplier-Key": supplier_key},
    )["token"]


def build_record(source, brand, brand_slug, model, model_key, photos, vin_info):
    row = source["row"]
    vin = (vin_info or {}).get("vin", "")
    notes = [
        "汽修宝批量导入。",
        "里程数为系统默认 99,999 km，仅为占位，不代表真实里程。",
        f"原始车型: {source['trim']}",
    ]
    if source.get("description"):
        notes.append(f"原始说明: {source['description']}")
    if (vin_info or {}).get("confidence"):
        notes.append(f"VIN OCR confidence: {vin_info['confidence']}")
    stock_id = f"QXB{row:04d}"
    title = f"{brand} {model} Half Cut"
    slug = "-".join(filter(bool, [brand_slug, slugify(model), str(source.get("year") or ""), "half-cut", stock_id.lower()]))
    return {
        "stockId": stock_id,
        "vin": vin,
        "decodeMethod": "QXB OCR" if vin else "Manual Entry",
        "decodeConfidence": (vin_info or {}).get("confidence") or None,
        "vehicleCondition": "Half Cut",
        "vehicleCategory": "passenger",
        "truckPartType": "",
        "machineryType": "",
        "brand": brand,
        "brandSlug": brand_slug,
        "model": model,
        "year": source.get("year") or None,
        "engineCode": "",
        "transmissionCode": "",
        "drivetrain": "2WD",
        "mileage": "99,999 km",
        "priceUsd": None,
        "origin": "China",
        "status": "Available",
        "title": title,
        "slug": slug,
        "photos": photos,
        "video": None,
        "videoUrl": "",
        "includedParts": ["Engine & gearbox assembly", "Front clip"],
        "shortDescription": f"{source.get('year') or ''} {brand} {model} — QXB sourced listing via AsiaPower.".strip(),
        "supplierVerified": True,
        "supplierName": "汽修宝",
        "supplierPhone": "16638801930",
        "supplierCity": "",
        "submissionId": f"QXB-{row:04d}",
        "approvedAt": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "notes": "\n".join(notes),
        "qxb": {
            "row": row,
            "brandCn": source["brand_cn"],
            "modelKey": model_key,
            "description": source.get("description") or "",
        },
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", type=pathlib.Path, default=DEFAULT_XLSX)
    parser.add_argument("--manifest", type=pathlib.Path, default=DEFAULT_MANIFEST)
    parser.add_argument("--vin-csv", type=pathlib.Path, default=DEFAULT_VIN)
    parser.add_argument("--state", type=pathlib.Path, default=DEFAULT_STATE)
    parser.add_argument("--out", type=pathlib.Path, default=DEFAULT_OUT)
    parser.add_argument("--supplier-key", default=os.environ.get("SUPPLIER_UPLOAD_KEY", ""))
    parser.add_argument("--limit", type=int)
    parser.add_argument("--max-photos", type=int, default=15)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument(
        "--skip-knowledge",
        action="store_true",
        help="Do not write upload harvest into shared knowledge base",
    )
    parser.add_argument(
        "--remember-knowledge",
        action="store_true",
        help="Append batch ingest summary to agent memory (inventory category)",
    )
    args = parser.parse_args()

    if not args.supplier_key and not args.dry_run:
        raise SystemExit("SUPPLIER_UPLOAD_KEY required")

    cn_to_en, cn_to_slug = load_brand_maps()
    manifest = load_manifest(args.manifest)
    vins = load_vins(args.vin_csv)
    upload_state = load_json(args.state, {"uploads": {}})
    records = []
    token = None

    for idx, source in enumerate(load_rows(args.xlsx), start=1):
        if args.limit and len(records) >= args.limit:
            break
        row = source["row"]
        local_photos = manifest.get(row, [])
        if len(local_photos) < 3:
            continue
        brand = cn_to_en.get(source["brand_cn"], source["brand_cn"])
        brand_slug = cn_to_slug.get(source["brand_cn"], slugify(brand))
        model_key, latin = parse_model(source["brand_cn"], source["trim"])
        model = model_english(brand_slug, model_key, latin)
        picks = selected_photos(row, local_photos, vins.get(row), args.max_photos)
        if len(picks) < 3:
            continue
        uploaded = []
        for pick in picks:
            path = pathlib.Path(pick["path"])
            cache_key = hashlib.sha1(f"{path}:{pick['label']}".encode()).hexdigest()
            if cache_key not in upload_state["uploads"] and not args.dry_run:
                if token is None:
                    token = get_upload_token(args.supplier_key)
                upload_state["uploads"][cache_key] = upload_file(path, pick["label"], args.supplier_key, token)
                save_json(args.state, upload_state)
            uploaded.append(upload_state["uploads"].get(cache_key) or {
                "label": pick["label"],
                "url": str(path),
                "fileName": path.name,
                "mimeType": mimetypes.guess_type(str(path))[0] or "image/jpeg",
                "size": path.stat().st_size,
            })
        records.append(build_record(source, brand, brand_slug, model, model_key, uploaded, vins.get(row)))
        if len(records) % 25 == 0:
            print(f"records={len(records)} uploads_cached={len(upload_state['uploads'])}", flush=True)

    save_json(args.out, records)

    knowledge_summary = None
    if records and not args.skip_knowledge:
        out_ref = str(args.out.relative_to(ROOT)) if args.out.is_relative_to(ROOT) else str(args.out)
        knowledge_summary = knowledge_ingest.ingest_batch(
            records,
            source="qxb_upload",
            source_ref_prefix=out_ref,
            dry_run=args.dry_run,
        )
        if args.remember_knowledge and not args.dry_run:
            knowledge_ingest.remember_batch_summary(knowledge_summary)

    print(json.dumps({
        "records": len(records),
        "uploadCache": len(upload_state["uploads"]),
        "out": str(args.out),
        "state": str(args.state),
        "dryRun": args.dry_run,
        "knowledge": knowledge_summary,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
