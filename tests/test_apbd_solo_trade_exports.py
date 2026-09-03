from __future__ import annotations

import csv
import json
from pathlib import Path

import pytest

from agents.apbd.solo_trade.exports import export_campaign


@pytest.fixture()
def campaign() -> dict:
    return {
        "campaign_id": "solo-fixture",
        "leads": [
            {
                "lead_id": "lead-1",
                "company": "Tema Diesel Works",
                "country": "Ghana",
                "city": "Tema",
                "website": "https://example.test",
                "public_email": "buyer@example.test",
                "business_type": "Importer",
                "status": "approval_pending",
                "score": {
                    "overall_score": 76.5,
                    "grade": "B",
                    "confidence": 71,
                    "source_urls": ["https://example.test/about"],
                },
                "activities": [{"event": "approval_requested", "at": "2026-09-02T12:00:00Z", "source": "test", "evidence_ref": "draft-1", "note": ""}],
            }
        ],
    }


def test_json_csv_and_xlsx_exports(tmp_path: Path, campaign: dict) -> None:
    outputs = export_campaign(campaign, tmp_path, "json")
    assert json.loads(Path(outputs["json"]).read_text())["campaign_id"] == "solo-fixture"
    outputs = export_campaign(campaign, tmp_path, "csv")
    with Path(outputs["csv"]).open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    assert rows[0]["company"] == "Tema Diesel Works"
    pytest.importorskip("openpyxl")
    outputs = export_campaign(campaign, tmp_path, "xlsx")
    from openpyxl import load_workbook

    workbook = load_workbook(outputs["xlsx"], data_only=False)
    assert workbook.sheetnames == ["Dashboard", "Leads", "Activities"]
    assert workbook["Leads"]["A2"].value == "lead-1"


def test_pdf_export_when_reportlab_is_available(tmp_path: Path, campaign: dict) -> None:
    pytest.importorskip("reportlab")
    outputs = export_campaign(campaign, tmp_path, "pdf")
    data = Path(outputs["pdf"]).read_bytes()
    assert data.startswith(b"%PDF")
    assert len(data) > 1000
