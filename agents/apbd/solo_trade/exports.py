"""Campaign exports for JSON, CSV, Excel, and PDF."""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from agents.apbd.solo_trade.crm import build_dashboard

LEAD_COLUMNS = (
    "lead_id",
    "company",
    "country",
    "city",
    "website",
    "public_email",
    "business_type",
    "status",
    "overall_score",
    "grade",
    "confidence",
    "source_urls",
)


def _lead_row(lead: dict[str, Any]) -> dict[str, Any]:
    score = lead.get("score") or {}
    urls = score.get("source_urls") or lead.get("source_urls") or [lead.get("source_url")]
    return {
        "lead_id": lead.get("lead_id") or lead.get("id") or "",
        "company": lead.get("company") or lead.get("display_name") or "",
        "country": lead.get("country") or "",
        "city": lead.get("city") or "",
        "website": lead.get("website") or "",
        "public_email": lead.get("public_email") or lead.get("email") or "",
        "business_type": lead.get("business_type") or lead.get("industry") or "",
        "status": lead.get("status") or "new",
        "overall_score": score.get("overall_score"),
        "grade": score.get("grade") or "",
        "confidence": score.get("confidence"),
        "source_urls": " | ".join(str(url) for url in urls if url),
    }


def export_json(campaign: dict[str, Any], path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(campaign, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return path


def export_csv(campaign: dict[str, Any], path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=LEAD_COLUMNS)
        writer.writeheader()
        for lead in campaign.get("leads") or []:
            writer.writerow(_lead_row(lead))
    return path


def export_xlsx(campaign: dict[str, Any], path: Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    dashboard = workbook.active
    dashboard.title = "Dashboard"
    summary = build_dashboard(campaign)
    dashboard.append(["Metric", "Value"])
    for key in ("campaign_id", "lead_count", "average_score", "sent_count", "reply_count", "reply_rate", "approval_pending"):
        dashboard.append([key, summary.get(key)])
    dashboard["A10"] = "Status"
    dashboard["B10"] = "Count"
    for index, (status, count) in enumerate(sorted(summary["status_counts"].items()), start=11):
        dashboard.cell(index, 1, status)
        dashboard.cell(index, 2, count)

    leads_sheet = workbook.create_sheet("Leads")
    leads_sheet.append(list(LEAD_COLUMNS))
    for lead in campaign.get("leads") or []:
        row = _lead_row(lead)
        leads_sheet.append([row[column] for column in LEAD_COLUMNS])
    statuses = "new,researched,draft_ready,approval_pending,approved,sent,opened,replied,qualified,disqualified,rejected"
    validation = DataValidation(type="list", formula1=f'"{statuses}"', allow_blank=False)
    leads_sheet.add_data_validation(validation)
    validation.add(f"H2:H{max(2, leads_sheet.max_row)}")
    if leads_sheet.max_row >= 2:
        leads_sheet.conditional_formatting.add(
            f"I2:I{leads_sheet.max_row}",
            ColorScaleRule(start_type="min", start_color="F3B8B5", mid_type="percentile", mid_value=50, mid_color="F8E7A7", end_type="max", end_color="B9DFC6"),
        )

    activities = workbook.create_sheet("Activities")
    activity_headers = ["lead_id", "company", "event", "at", "source", "evidence_ref", "note"]
    activities.append(activity_headers)
    for lead in campaign.get("leads") or []:
        for activity in lead.get("activities") or []:
            activities.append(
                [
                    lead.get("lead_id") or lead.get("id") or "",
                    lead.get("company") or lead.get("display_name") or "",
                    activity.get("event") or "",
                    activity.get("at") or "",
                    activity.get("source") or "",
                    activity.get("evidence_ref") or "",
                    activity.get("note") or "",
                ]
            )
    header_fill = PatternFill("solid", fgColor="153B35")
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")
        for column_cells in sheet.columns:
            values = [str(cell.value or "") for cell in column_cells[:100]]
            width = min(48, max(12, max(len(value) for value in values) + 2))
            sheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)
    return path


def export_pdf(campaign: dict[str, Any], path: Path) -> Path:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    path.parent.mkdir(parents=True, exist_ok=True)
    styles = getSampleStyleSheet()
    dashboard = build_dashboard(campaign)
    story: list[Any] = [
        Paragraph("APBD Solo Trade Campaign", styles["Title"]),
        Paragraph(f"Campaign: {campaign.get('campaign_id', '')}", styles["Normal"]),
        Paragraph("Evidence estimates only. External sending requires separate approval.", styles["Normal"]),
        Spacer(1, 6 * mm),
    ]
    metrics = [["Leads", "Average score", "Sent", "Replies", "Approval pending"], [
        dashboard["lead_count"],
        dashboard["average_score"] if dashboard["average_score"] is not None else "N/A",
        dashboard["sent_count"],
        dashboard["reply_count"],
        dashboard["approval_pending"],
    ]]
    metric_table = Table(metrics, repeatRows=1)
    metric_table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#153B35")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#B8C5BF")), ("PADDING", (0, 0), (-1, -1), 6)]))
    story.extend([metric_table, Spacer(1, 6 * mm)])
    rows = [["Company", "Market", "Status", "Score", "Grade", "Evidence URL"]]
    for lead in campaign.get("leads") or []:
        row = _lead_row(lead)
        rows.append([row["company"][:34], f"{row['city']} {row['country']}"[:24], row["status"], row["overall_score"] if row["overall_score"] is not None else "N/A", row["grade"], row["source_urls"][:55]])
    table = Table(rows, repeatRows=1, colWidths=[48 * mm, 34 * mm, 28 * mm, 18 * mm, 14 * mm, 80 * mm])
    table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#153B35")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("FONTSIZE", (0, 0), (-1, -1), 7), ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#B8C5BF")), ("VALIGN", (0, 0), (-1, -1), "TOP"), ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F6F3")])]))
    story.append(table)
    document = SimpleDocTemplate(str(path), pagesize=landscape(A4), leftMargin=12 * mm, rightMargin=12 * mm, topMargin=12 * mm, bottomMargin=12 * mm)
    document.build(story)
    return path


def export_campaign(campaign: dict[str, Any], output_dir: Path, fmt: str = "all") -> dict[str, str]:
    selected = {"json", "csv", "xlsx", "pdf"} if fmt == "all" else {fmt.casefold()}
    unsupported = selected - {"json", "csv", "xlsx", "pdf"}
    if unsupported:
        raise ValueError(f"Unsupported export format: {', '.join(sorted(unsupported))}")
    campaign_id = str(campaign.get("campaign_id") or "campaign")
    outputs: dict[str, str] = {}
    functions = {"json": export_json, "csv": export_csv, "xlsx": export_xlsx, "pdf": export_pdf}
    for extension in ("json", "csv", "xlsx", "pdf"):
        if extension in selected:
            path = output_dir / f"{campaign_id}.{extension}"
            functions[extension](campaign, path)
            outputs[extension] = str(path)
    return outputs
