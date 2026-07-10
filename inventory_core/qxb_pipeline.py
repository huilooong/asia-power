"""QXB (汽修宝) upload pipeline — one-row-at-a-time workflow for APInventory."""

from __future__ import annotations

import csv
import hashlib
import html
import io
import json
import mimetypes
import os
import re
import shutil
import socket
import subprocess
import sys
import time
import urllib.error
import urllib.request
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent

MFR_PREFIX = ["北京现代", "东风", "一汽", "广汽", "上汽", "长安", "华晨", "进口"]
YEAR_RE = re.compile(r"(\d{4})款")
VIN_STRICT = re.compile(r"^[A-HJ-NPR-Z][A-HJ-NPR-Z0-9]{16}$")
DISP_RE = re.compile(r"^\d\.\d[TL]?$|^\d\.\d$")
DISP_TOKEN_RE = re.compile(r"^\d+\.\d+[TL]?$", re.I)
TRIM_CONFIG_WORDS = frozenset(
    "手动 自动 舒适型 豪华型 标准型 时尚型 基本型 精英型 尊贵型 旗舰型 领先型 尊享型 豪华领航版 两驱 四驱".split()
)
STOP = set("MT AT CVT DCT AMT DSG 三厢 两厢 旅行版 旅行 掀背 双擎 混动".split())

from inventory_core.qxb_photo_pick import pick_photo_slots
from inventory_core.qxb_price_estimate import _model_match, estimate_half_cut_price_usd

PHOTO_LABELS = [
    "Front view",
    "Rear view",
    "Engine bay",
    "VIN / chassis plate",
    "Interior",
]

DEFAULT_UPLOAD_MAX_PHOTOS = int(os.environ.get("QXB_UPLOAD_MAX_PHOTOS", "40"))


def bulk_upload_mode() -> bool:
    """CEO bulk pass: all manifest photos, no 子龙 slot training."""
    return os.environ.get("QXB_BULK_ALL_PHOTOS", "").strip().lower() in ("1", "true", "yes")


def batch_skip_vin_decode() -> bool:
    """Batch re-check: avoid 60/hr decode quota — use trim/nameplate fallback."""
    return os.environ.get("QXB_BATCH_SKIP_DECODE", "").strip().lower() in ("1", "true", "yes")


def _effective_max_photos(fields: dict[str, Any], max_photos: int) -> int:
    if bulk_upload_mode():
        return max(len(fields.get("picks") or []), max_photos)
    return max_photos


def all_manifest_photo_picks(photos: list[dict]) -> tuple[list[dict[str, str]], dict[str, Any]]:
    """Use every downloaded photo (sorted by index) — no 子龙 slot picking."""
    if not photos:
        return [], {"method": "bulk_all_photos", "confidence": "bulk", "count": 0}

    def _idx(p: dict) -> int:
        try:
            return int(p.get("image_index") or 0)
        except (TypeError, ValueError):
            return 0

    picks: list[dict[str, str]] = []
    for p in sorted(photos, key=_idx):
        path = str(p.get("local_path") or p.get("path") or "").strip()
        if not path or not Path(path).is_file():
            continue
        idx = _idx(p)
        label = f"Photo {idx:02d}" if idx else Path(path).stem
        picks.append({"label": label, "path": path})
    return picks, {"method": "bulk_all_photos", "confidence": "bulk", "count": len(picks)}

ROW_STATUSES = frozenset({
    "pending", "inspected", "prepared", "uploaded", "knowledge",
    "pending_review", "blocked", "skipped", "parked",
})

PARK_CATEGORIES: dict[str, str] = {
    "submit_ghost": "本地已标记提交，Admin 无记录 — 待补推",
    "no_vin": "无 VIN OCR — 稍后处理",
    "album_incomplete": "相册不全 — 待补图或 CEO 标注",
    "awaiting_ceo": "Admin Pending — 待 CEO 审核/操作",
    "photo_uncertain": "识图未确认 — 待 CEO 标注后再传",
}

PIPELINE_VERSION = "1.4.0"

# Align QXB slot labels with supplier portal / admin review UI.
PHOTO_LABEL_TO_REVIEW = {
    "Front view": "Vehicle Front",
    "Rear view": "Vehicle Rear",
    "Engine bay": "Engine",
    "VIN / chassis plate": "VIN Plate",
    "Interior": "Interior",
}


@dataclass
class QxbPaths:
    root: Path = field(default_factory=lambda: ROOT)
    xlsx: Path | None = None
    manifest: Path | None = None
    vin_csv: Path | None = None
    upload_state: Path | None = None
    approved_out: Path | None = None
    agent_queue: Path | None = None
    model_dict: Path | None = None
    brand_dict: Path | None = None
    api_base: str = "https://asia-power.com"

    def __post_init__(self) -> None:
        if self.xlsx is None:
            self.xlsx = self._default_xlsx()
        self.manifest = self.manifest or self.root / "data/qxb-photos/manifest.csv"
        self.vin_csv = self.vin_csv or self.root / "reports/qxb-vin-ocr-results.csv"
        self.upload_state = self.upload_state or self.root / "reports/qxb-site-upload-state.json"
        self.approved_out = self.approved_out or self.root / "reports/qxb-approved-import.json"
        self.agent_queue = self.agent_queue or self.root / "reports/qxb-agent-queue.json"
        self.model_dict = self.model_dict or self.root / "data/knowledge-base/model-dictionary.json"
        self.brand_dict = self.brand_dict or self.root / "data/knowledge-base/brand-dictionary.json"

    def _default_xlsx(self) -> Path:
        active = self.root / "data/qxb-vehicles-active.xlsx"
        if active.is_file():
            return active
        downloads_new = Path("/Users/longhui/Downloads/汽修宝车辆数据_060708(1).xlsx")
        if downloads_new.is_file():
            return downloads_new
        downloads = Path("/Users/longhui/Downloads/汽修宝车辆数据_2606291619.xlsx")
        if downloads.is_file():
            return downloads
        return self.root / "data/qxb-vehicles.xlsx"


def reconfigure_paths(**kwargs: Any) -> QxbPaths:
    """Build paths for tests (returns new QxbPaths instance)."""
    return QxbPaths(**kwargs)


def _now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")


def _iso_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def load_json(path: Path, fallback: Any) -> Any:
    if not path.is_file():
        return fallback
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return fallback


def save_json_atomic(path: Path, data: Any) -> None:
    """Write JSON atomically; unique tmp + retry avoids parallel batch races."""
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = json.dumps(data, ensure_ascii=False, indent=2) + "\n"
    last_err: OSError | None = None
    for attempt in range(3):
        tmp = path.with_name(f"{path.name}.{os.getpid()}.{attempt}.tmp")
        try:
            tmp.write_text(payload, encoding="utf-8")
            tmp.replace(path)
            return
        except OSError as exc:
            last_err = exc
            if tmp.is_file():
                try:
                    tmp.unlink()
                except OSError:
                    pass
            time.sleep(0.05 * (attempt + 1))
    if last_err:
        raise last_err


def slugify(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", str(value or "").lower()).strip("-")


def _is_displacement_token(token: str) -> bool:
    t = str(token or "").strip()
    return bool(DISP_RE.match(t) or DISP_TOKEN_RE.match(t))


def normalize_decode_model_name(
    decode_model: str,
    brand: str = "",
    brand_cn: str = "",
) -> str:
    """Prefer VIN-decode model over Excel trim heuristics (e.g. 福克斯 not 2012 1 8L)."""
    name = str(decode_model or "").strip()
    if not name:
        return name
    for prefix in (brand, brand_cn):
        p = str(prefix or "").strip()
        if p and name.startswith(p):
            name = name[len(p):].strip()
    return name or str(decode_model or "").strip()


def apply_decode_listing_identity(fields: dict[str, Any]) -> None:
    """When VIN decode succeeds, use decode brand/model for listing fields."""
    decode = fields.get("vin_decode") or {}
    if not decode.get("ok"):
        return
    source = fields.get("source") or {}
    brand_cn = str(source.get("brand_cn") or "")
    decode_brand = str(decode.get("brand") or "").strip()
    if decode_brand:
        fields["brand"] = decode_brand
        fields["brand_slug"] = slugify(decode_brand)
    decode_model = normalize_decode_model_name(
        str(decode.get("model") or ""),
        fields.get("brand") or decode_brand,
        brand_cn,
    )
    if decode_model:
        fields["model"] = decode_model
        fields["model_key"] = decode_model
    if decode.get("year") and not source.get("year"):
        source["year"] = decode.get("year")


def parse_model(brand: str, trim: str) -> tuple[str, str | None]:
    """Extract model name from 汽修宝 trim like '2012款 三厢经典 1.8L 手动时尚型'."""
    s = YEAR_RE.sub("", str(trim or "").strip(), count=1).strip()
    tokens = [t for t in s.split() if t]
    if not tokens:
        return trim or "", None

    disp_idx = next((i for i, t in enumerate(tokens) if _is_displacement_token(t)), None)
    model_tokens: list[str] = []
    if disp_idx is not None:
        model_tokens = [t for t in tokens[:disp_idx] if t not in STOP]
        if not model_tokens:
            for t in tokens[disp_idx + 1 :]:
                if t in STOP or t in TRIM_CONFIG_WORDS:
                    break
                model_tokens.append(t)
    else:
        for t in tokens:
            if _is_displacement_token(t) or t in STOP or t in TRIM_CONFIG_WORDS:
                break
            model_tokens.append(t)

    raw = "".join(model_tokens).strip()
    if not raw:
        raw = s
    cjk = "".join(re.findall(r"[\u4e00-\u9fff]+", raw))
    latin = " ".join(re.findall(r"[A-Za-z0-9\-]+", raw))
    if cjk and latin:
        return cjk, latin
    if latin and not cjk:
        return latin, latin
    return raw, None


def load_rows(xlsx: Path) -> list[dict[str, Any]]:
    if not xlsx.is_file():
        return []
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active
    rows: list[dict[str, Any]] = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        brand = str(row[0] or "").strip()
        trim = str(row[1] or "").strip()
        desc = str(row[2] or "").strip()
        if not brand or not trim:
            continue
        year_match = YEAR_RE.search(trim)
        rows.append({
            "row": row_num,
            "brand_cn": brand,
            "trim": trim,
            "description": desc,
            "year": int(year_match.group(1)) if year_match else None,
        })
    return rows


def load_brand_maps(brand_dict_path: Path) -> tuple[dict[str, str], dict[str, str]]:
    brand_dict = load_json(brand_dict_path, {})
    cn_to_en: dict[str, str] = {}
    cn_to_slug: dict[str, str] = {}
    for cn, value in brand_dict.items():
        en = value.get("english") or cn
        cn_to_en[cn] = en
        cn_to_slug[cn] = slugify(en)
    return cn_to_en, cn_to_slug


def model_english(model_dict_path: Path, brand_slug: str, model_key: str, fallback: str | None) -> str:
    md = load_json(model_dict_path, {})
    rec = md.get(brand_slug, {}).get(model_key) or {}
    return rec.get("english") or fallback or model_key


def load_manifest(path: Path) -> dict[int, list[dict[str, str]]]:
    by_row: dict[int, list[dict[str, str]]] = defaultdict(list)
    if not path.is_file():
        return by_row
    with path.open(encoding="utf-8") as fh:
        for rec in csv.DictReader(fh):
            if rec.get("status") not in {"downloaded", "exists"}:
                continue
            local = Path(rec["local_path"])
            if not local.is_file() or local.stat().st_size <= 0:
                continue
            by_row[int(rec["row"])].append(rec)
    for items in by_row.values():
        items.sort(key=lambda r: int(r["image_index"]))
    return by_row


def normalize_vin_strict(vin: str) -> str:
    """Match admin HalfCutVin rules; fix common OCR O/0, I/1, Q/0 swaps."""
    raw = str(vin or "").strip().upper()
    if VIN_STRICT.match(raw):
        return raw
    for src, dst in (("O", "0"), ("I", "1"), ("Q", "0")):
        if src not in raw:
            continue
        candidate = raw.replace(src, dst)
        if VIN_STRICT.match(candidate):
            return candidate
    return raw


def format_duplicate_vin_message(dup: dict[str, Any] | None) -> str:
    """CEO-facing Chinese message when the same VIN was already uploaded."""
    if not dup:
        return ""
    detail: list[str] = []
    if dup.get("stockId"):
        detail.append(str(dup["stockId"]))
    if dup.get("row") is not None:
        detail.append(f"row {dup['row']}")
    if dup.get("submissionId"):
        detail.append(str(dup["submissionId"]))
    if detail:
        return f"此底盘号已上传过，禁止重复上传（已有 {' / '.join(detail)}）"
    return "此底盘号已上传过，禁止重复上传"


def _vin_row_from_stock(stock_id: str) -> int | None:
    sid = str(stock_id or "").strip().upper()
    if sid.startswith("QXB") and len(sid) >= 7:
        try:
            return int(sid[3:])
        except ValueError:
            return None
    return None


def _load_review_upload_queue_items(root: Path) -> list[dict[str, Any]]:
    path = root / "reports/qxb-review-upload-queue.json"
    if not path.is_file():
        return []
    try:
        data = load_json(path, {"items": []})
    except Exception:
        return []
    items = data.get("items") if isinstance(data, dict) else None
    return list(items) if isinstance(items, list) else []


def find_duplicate_vin(
    ctx: PipelineContext,
    vin: str,
    *,
    exclude_row: int | None = None,
    server_state: dict[str, Any] | None = None,
    upload_queue_items: list[dict[str, Any]] | None = None,
) -> dict[str, Any] | None:
    """Return first duplicate occurrence for normalized 17-char VIN (different row/stockId)."""
    vin_norm = normalize_vin_strict(vin)
    if not vin_norm or not VIN_STRICT.match(vin_norm):
        return None

    exclude_stock = f"QXB{exclude_row:04d}" if exclude_row is not None else None

    def _skip(row_n: int | None, stock_id: str | None = None) -> bool:
        if exclude_row is not None and row_n == exclude_row:
            return True
        if exclude_stock and stock_id and str(stock_id).upper() == exclude_stock:
            return True
        return False

    for row_n, info in ctx.vins.items():
        if info.get("vin") != vin_norm:
            continue
        stock_id = f"QXB{row_n:04d}"
        if _skip(row_n, stock_id):
            continue
        return {"vin": vin_norm, "row": row_n, "stockId": stock_id, "source": "vin_csv"}

    for rec in ctx.approved:
        if normalize_vin_strict(rec.get("vin") or "") != vin_norm:
            continue
        stock_id = str(rec.get("stockId") or "")
        row_n = (rec.get("qxb") or {}).get("row")
        if row_n is None:
            row_n = _vin_row_from_stock(stock_id)
        if _skip(row_n, stock_id):
            continue
        return {
            "vin": vin_norm,
            "row": row_n,
            "stockId": stock_id,
            "source": "approved_local",
            "submissionId": rec.get("submissionId"),
        }

    for key, entry in (ctx.queue.get("rows") or {}).items():
        status = str(entry.get("status") or "")
        if status not in ("pending_review", "uploaded", "knowledge"):
            continue
        row_n = int(key)
        stock_id = f"QXB{row_n:04d}"
        if _skip(row_n, stock_id):
            continue
        row_vin = normalize_vin_strict((ctx.vins.get(row_n) or {}).get("vin") or "")
        if not row_vin:
            for rec in ctx.approved:
                if rec.get("stockId") == stock_id:
                    row_vin = normalize_vin_strict(rec.get("vin") or "")
                    break
        if row_vin != vin_norm:
            continue
        return {
            "vin": vin_norm,
            "row": row_n,
            "stockId": stock_id,
            "source": "queue",
            "submissionId": entry.get("submissionId"),
            "status": status,
        }

    queue_items = (
        upload_queue_items
        if upload_queue_items is not None
        else _load_review_upload_queue_items(ctx.paths.root)
    )
    for item in queue_items:
        if str(item.get("status") or "") not in ("pending", "processing", "done"):
            continue
        row_n = item.get("row")
        if row_n is None:
            continue
        try:
            row_n = int(row_n)
        except (TypeError, ValueError):
            continue
        stock_id = str(item.get("stockId") or f"QXB{row_n:04d}")
        if _skip(row_n, stock_id):
            continue
        row_vin = normalize_vin_strict((ctx.vins.get(row_n) or {}).get("vin") or "")
        if row_vin != vin_norm:
            continue
        return {
            "vin": vin_norm,
            "row": row_n,
            "stockId": stock_id,
            "source": "upload_queue",
            "status": item.get("status"),
        }

    if server_state:
        for sub in server_state.get("submissions") or []:
            if normalize_vin_strict(sub.get("vin") or "") != vin_norm:
                continue
            review = str(sub.get("reviewStatus") or "pending")
            if review == "rejected":
                continue
            stock_id = str(sub.get("qxbStockId") or sub.get("stockId") or "")
            row_n = (sub.get("qxb") or {}).get("row")
            if row_n is None:
                row_n = _vin_row_from_stock(stock_id)
            if _skip(row_n, stock_id or None):
                continue
            return {
                "vin": vin_norm,
                "stockId": stock_id or None,
                "row": row_n,
                "source": "server_pending" if review == "pending" else f"server_{review}",
                "submissionId": sub.get("submissionId"),
            }
        for item in server_state.get("approved") or []:
            if normalize_vin_strict(item.get("vin") or "") != vin_norm:
                continue
            stock_id = str(item.get("stockId") or item.get("qxbStockId") or "")
            row_n = (item.get("qxb") or {}).get("row")
            if row_n is None:
                row_n = _vin_row_from_stock(stock_id)
            if _skip(row_n, stock_id or None):
                continue
            return {
                "vin": vin_norm,
                "stockId": stock_id or None,
                "row": row_n,
                "source": "server_approved",
                "submissionId": item.get("submissionId"),
            }

    return None


GARBAGE_VIN_TERMS = re.compile(
    r"(CVT|ATGLS|MTGL|SKCVT|AUT|MANUAL|GLS|PREMIUM|COMFORT|LUXURY|NAVI|HYBRID|DCT|DSG|AMT|"
    r"REANP|STSFE|ERRU|BFRHP|SCUNH|PRM0F|ALBR0|RULER|NEREE|BANDST|PERERE|HSREME|SARNW|"
    r"SKCVT2WD|A82010520|A546WSA7|A290E040|A20100804|A13AEF4A|A2NYAYW2|A4RP2PSS|A27W4SLE|"
    r"A73ADPL4|A102RFEY|A2F3PMBN|AABE4KH2|A4GGHBAY|A3VEHBVP|A82010520|A80SY286|A4NSGEPS|"
    r"ARR180R7|HSOBLLO|PERNTSZ|USRCCRDA|FTLBP8DE|S47NYNUU|RSSSRNQS|Y25BPS3Y|NEHHAAGH|"
    r"S7297ALAT|BARBPRCA|GRRATBRA|BANDST0E|ABZCYPBZ|ANSSEANK|AYYB5TRF|ABKRJ32B|A4TS4WY2|"
    r"ARHEFC25|AYAPTRYB|PERNTSZB|A82010520|A546WSA7P|A290E0407|A13AEF4AS|A2NYAYW2L)",
    re.I,
)


def is_garbage_ocr_vin(vin: str) -> bool:
    """Reject OCR misreads that look like 17 chars but embed trim words / noise."""
    raw = str(vin or "").strip().upper()
    if not raw or len(raw) != 17:
        return True
    if not VIN_STRICT.match(raw):
        return True
    if GARBAGE_VIN_TERMS.search(raw):
        return True
    # VIN check digit position 9 is usually numeric; all-alpha middle blocks are rare
    if raw[8].isalpha() and sum(ch.isalpha() for ch in raw[3:8]) >= 4:
        if re.search(r"(SSSS|AAAA|BBBB|CCCC|EEEE|RRRR|NNNN|LLLL|PPPP)", raw):
            return True
    return False


def clear_vin_csv_row(path: Path, row: int) -> None:
    """Remove a row from VIN OCR CSV so ensure_row_vin can rescan."""
    rows = [rec for rec in _read_vin_csv_rows(path) if rec.get("row") != str(row)]
    _write_vin_csv_rows(path, rows)


def load_vins(path: Path) -> dict[int, dict[str, str]]:
    vins: dict[int, dict[str, str]] = {}
    for rec in _read_vin_csv_rows(path):
        vin = normalize_vin_strict(str(rec.get("vin") or ""))
        if not vin or not VIN_STRICT.match(vin) or is_garbage_ocr_vin(vin):
            continue
        vins[int(rec["row"])] = {
            "vin": vin,
            "image_path": rec.get("image_path") or "",
            "confidence": rec.get("confidence") or "",
            "engine_code": str(rec.get("engine_code") or "").strip().upper(),
            "transmission_code": str(rec.get("transmission_code") or "").strip().upper(),
        }
    return vins


VIN_CSV_FIELDS = ["row", "model", "vin", "image_path", "confidence", "engine_code", "transmission_code"]


def _read_vin_csv_rows(path: Path) -> list[dict[str, str]]:
    """Load VIN OCR CSV; drop corrupt lines (invalid UTF-8 or non-numeric row)."""
    if not path.is_file():
        return []
    text = path.read_bytes().decode("utf-8", errors="replace")
    rows: list[dict[str, str]] = []
    for rec in csv.DictReader(io.StringIO(text)):
        row_key = str(rec.get("row") or "").strip()
        if not row_key.isdigit():
            continue
        rows.append({field: str(rec.get(field) or "") for field in VIN_CSV_FIELDS})
    return rows


def _write_vin_csv_rows(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    with tmp.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=VIN_CSV_FIELDS)
        writer.writeheader()
        writer.writerows(rows)
    tmp.replace(path)


def write_vin_csv_row(
    path: Path,
    row: int,
    model: str,
    vin: str,
    image_path: str,
    confidence: str,
    *,
    engine_code: str = "",
    transmission_code: str = "",
) -> None:
    """Upsert one row in reports/qxb-vin-ocr-results.csv."""
    rows = _read_vin_csv_rows(path)
    existing = next((rec for rec in rows if rec.get("row") == str(row)), {})
    entry = {
        "row": str(row),
        "model": model or existing.get("model") or "",
        "vin": vin,
        "image_path": image_path or existing.get("image_path") or "",
        "confidence": confidence or existing.get("confidence") or "",
        "engine_code": str(engine_code or existing.get("engine_code") or "").strip().upper(),
        "transmission_code": str(transmission_code or existing.get("transmission_code") or "").strip().upper(),
    }
    found = False
    for rec in rows:
        if rec.get("row") == str(row):
            rec.update(entry)
            found = True
            break
    if not found:
        rows.append(entry)
    _write_vin_csv_rows(path, rows)


def photo_scan_order(photos: list[dict]) -> list[int]:
    """CEO-trained VIN affinities first, then every album photo for OCR."""
    from inventory_core.qxb_photo_pick import load_learnings, vin_photo_scan_indices

    store = load_learnings()
    model = store.get("recognitionModel") or {}
    return vin_photo_scan_indices(photos, model=model)


def ocr_vin_from_photos(
    photos: list[dict],
    *,
    timeout: int = 12,
    thorough: bool = True,
) -> tuple[str, str, str]:
    """OCR album for VIN; priority indices (3 / 17–19) first, then full scan."""
    from importlib.machinery import SourceFileLoader
    from inventory_core.qxb_photo_pick import QXB_VIN_SCAN_PRIORS

    ocr = SourceFileLoader("ocr", str(ROOT / "scripts/ocr-qxb-vins.py")).load_module()
    by_idx = {int(p["image_index"]): p for p in photos}

    def _scan(indices: list[int]) -> list[dict[str, Any]]:
        hits: list[dict[str, Any]] = []
        for idx in indices:
            item = by_idx.get(idx)
            if not item:
                continue
            path = Path(str(item.get("local_path") or ""))
            if not path.is_file():
                continue
            for hit in ocr.candidates_for(path, thorough=thorough, timeout=timeout):
                hits.append({**hit, "path": str(path), "idx": idx})
        return hits

    scan_order = photo_scan_order(photos)
    priority_set = set(QXB_VIN_SCAN_PRIORS)
    priority_order = [i for i in scan_order if i in priority_set]
    rest_order = [i for i in scan_order if i not in priority_set]

    row_hits = _scan(priority_order)
    strict = [h for h in row_hits if h.get("strict")]
    vin, conf = ocr.pick_best(strict or row_hits)
    if vin:
        plate_path = next((h["path"] for h in row_hits if h.get("vin") == vin), "")
        return vin, conf, plate_path

    row_hits.extend(_scan(rest_order))
    strict_all = [h for h in row_hits if h.get("strict")]
    vin, conf = ocr.pick_best(strict_all)
    if vin:
        plate_path = next((h["path"] for h in row_hits if h.get("vin") == vin), "")
        return vin, conf, plate_path
    return "", "none", ""


def ensure_row_vin(ctx: PipelineContext, row: int, *, force_rescan: bool = False) -> dict[str, str] | None:
    """Load VIN from CSV; if missing, OCR all photos from Excel-linked album and persist."""
    existing = ctx.vins.get(row)
    if existing and existing.get("vin") and not force_rescan:
        if not is_garbage_ocr_vin(existing["vin"]):
            return existing
        clear_vin_csv_row(ctx.paths.vin_csv, row)
        ctx.vins.pop(row, None)
        existing = None
    photos = ctx.manifest.get(row, [])
    if not photos:
        return None
    vin, conf, plate_path = ocr_vin_from_photos(photos, thorough=True, timeout=18)
    if not vin:
        return None
    vin_norm = normalize_vin_strict(vin)
    if not VIN_STRICT.match(vin_norm) or is_garbage_ocr_vin(vin_norm):
        return None
    model = photos[0].get("model") or (ctx.sources.get(row) or {}).get("trim") or ""
    confidence = f"album_ocr:{conf}"
    engine_code = ""
    if plate_path:
        try:
            from importlib.machinery import SourceFileLoader

            ocr = SourceFileLoader("ocr", str(ROOT / "scripts/ocr-qxb-vins.py")).load_module()
            engine_code = str((ocr.nameplate_powertrain_for(Path(plate_path)) or {}).get("engineCode") or "")
        except Exception:
            engine_code = ""
    write_vin_csv_row(
        ctx.paths.vin_csv,
        row,
        model,
        vin_norm,
        plate_path,
        confidence,
        engine_code=engine_code,
    )
    info = {
        "vin": vin_norm,
        "image_path": plate_path,
        "confidence": confidence,
        "engine_code": engine_code,
    }
    ctx.vins[row] = info
    return info


def auto_recover_row(ctx: PipelineContext, row: int) -> dict[str, Any]:
    """Try to fix common batch failures before inspect/upload."""
    actions: list[str] = []
    vin_info = ctx.vins.get(row)
    if vin_info and vin_info.get("vin") and is_garbage_ocr_vin(vin_info["vin"]):
        clear_vin_csv_row(ctx.paths.vin_csv, row)
        ctx.vins.pop(row, None)
        actions.append("cleared_garbage_vin")
        vin_info = None

    if not (vin_info or {}).get("vin"):
        found = ensure_row_vin(ctx, row, force_rescan=bool(actions))
        if found:
            actions.append(f"ocr_vin:{found['vin']}")
            ctx = load_context(ctx.paths)
        else:
            actions.append("ocr_no_vin")

    return {"row": row, "actions": actions}


def categorize_blocker(message: str | None) -> str:
    """Bucket blocker text for CEO batch reports."""
    msg = str(message or "")
    lower = msg.lower()
    if "ocr found no vin" in lower or "no vin" in lower:
        return "ocr_no_vin"
    if "invalid for admin approval" in lower or "ocr 0↔o" in lower:
        return "invalid_vin_ocr"
    if "rate_limited" in lower or "限流" in msg or "60次" in msg:
        return "vin_decode_rate_limit"
    if "qxb_unavailable" in lower or "quota" in lower or "9005" in msg or "汽修宝" in msg:
        return "vin_decode_qxb_quota"
    if "vin decode failed" in lower or "decode failed" in lower or "not_found" in lower:
        return "vin_decode_failed"
    if "此底盘号已上传过" in msg or "禁止重复上传" in msg:
        return "duplicate_vin"
    if "pending submission already exists" in lower:
        return "duplicate_vin_pending"
    if "utf-8" in lower or "codec can't decode" in lower:
        return "csv_corrupt_fixed"
    if "no such file" in lower and "upload-state" in lower:
        return "upload_state_race"
    if "502" in msg or "503" in msg or "bad gateway" in lower:
        return "http_transient"
    if "enginecode missing" in lower or "transmissioncode missing" in lower:
        return "powertrain_missing"
    return "other"


def selected_photos(
    photos: list[dict],
    vin_info: dict | None,
    *,
    row: int | None = None,
    max_photos: int | None = None,
) -> tuple[list[dict[str, str]], dict[str, Any]]:
    cap = max_photos if max_photos is not None else DEFAULT_UPLOAD_MAX_PHOTOS
    return pick_photo_slots(photos, vin_info, row=row, max_photos=cap)


VIN_DECODE_REASON_MESSAGES: dict[str, str] = {
    "rate_limited": "asia-power.com VIN 接口限流（60次/小时）— 稍后重试或走 trim 推断",
    "not_found": "汽修宝 VinDecoder 无此 VIN 记录",
    "invalid_vin": "VIN 格式无效（须 17 位）",
    "invalid_request": "VIN decode 请求格式错误",
    "qxb_unavailable": "汽修宝 VinDecoder 不可用（凭证/网络/配额）",
}


def format_vin_decode_error(body: dict[str, Any] | None) -> str:
    """Map /api/vin/decode JSON to a CEO-readable message (not generic 'decode failed')."""
    if not body:
        return "decode failed"
    reason = str(body.get("reason") or "").strip()
    if reason in VIN_DECODE_REASON_MESSAGES:
        msg = VIN_DECODE_REASON_MESSAGES[reason]
        detail = str(body.get("message") or body.get("error") or "").strip()
        return f"{msg} ({detail})" if detail else msg
    for key in ("error", "message"):
        val = str(body.get(key) or "").strip()
        if val:
            return val
    if reason:
        return reason
    return "decode failed"


VIN_DECODE_RETRY_REASONS = frozenset({"rate_limited", "qxb_unavailable"})
VIN_DECODE_RETRY_HTTP = frozenset({502, 503, 504})
DEFAULT_VIN_DECODE_MAX_RETRIES = int(os.environ.get("QXB_VIN_DECODE_RETRY_MAX", "3"))
DEFAULT_VIN_DECODE_BACKOFF_SEC = float(os.environ.get("QXB_VIN_DECODE_BACKOFF", "65"))

_VIN_DECODE_MEMORY: dict[str, dict[str, Any]] = {}


def _vin_decode_cache_path(root: Path | None = None) -> Path:
    return (root or ROOT) / "data/knowledge-base/vin-decode-cache.json"


def _approved_vin_facts_path(root: Path | None = None) -> Path:
    return (root or ROOT) / "data/knowledge-base/approved-vin-facts.json"


def _vin_decode_cache_tier(body: dict[str, Any]) -> int:
    """Higher tier wins on merge — API decode must not be overwritten by seed/fallback."""
    source = str(body.get("decodeSource") or "")
    method = str(body.get("decodeMethod") or "")
    if source in ("qxb_via_asia_power", "vin_cache_legacy") or method == "QXB":
        return 3
    if method in ("Trim+Catalog",):
        return 0
    if source in ("approved_import_seed", "approved_import", "approved_vin_facts"):
        return 2
    return 1


def _merge_vin_decode_payload(existing: dict[str, Any], incoming: dict[str, Any]) -> dict[str, Any]:
    if _vin_decode_cache_tier(incoming) < _vin_decode_cache_tier(existing):
        return dict(existing)
    merged = dict(incoming)
    if _vin_decode_cache_tier(existing) >= 3:
        for key in (
            "brand", "model", "year", "engineCode", "transmissionCode",
            "gearboxModel", "drivetrain", "fuelType", "displacement",
        ):
            if not str(merged.get(key) or "").strip() and str(existing.get(key) or "").strip():
                merged[key] = existing[key]
    return merged


def _map_qxb_raw_to_decode(
    vin: str,
    raw: dict[str, Any],
    *,
    source: str = "vin_cache_legacy",
) -> dict[str, Any] | None:
    """Map QXB VinDecoder raw JSON (vin-cache.json) to pipeline decode shape."""
    if not isinstance(raw, dict) or raw.get("number") != 200:
        return None
    models = (raw.get("result") or {}).get("models") or []
    if not models:
        return None
    model = models[0]
    engine = str(model.get("engine_code") or "").strip()
    trans_type = str(model.get("trans_type") or "").strip().upper()
    shift_num = str(model.get("shift_num") or "").strip()
    trans = f"{shift_num}{trans_type}" if shift_num.isdigit() and trans_type else trans_type
    trans_code = str((raw.get("result") or {}).get("trans_code") or "").strip()
    if trans_code and trans_code == engine:
        trans_code = ""
    brand = str(model.get("brand") or "").strip()
    series = str(model.get("series") or "").strip()
    model_name = str(model.get("model_name") or series or "").strip()
    year_raw = model.get("years")
    year = int(year_raw) if str(year_raw or "").isdigit() else None
    if not engine and not series and not model_name:
        return None
    return {
        "ok": True,
        "vin": vin,
        "brand": brand,
        "model": series or model_name,
        "year": year,
        "engineCode": engine,
        "transmissionCode": trans,
        "gearboxModel": trans_code,
        "drivetrain": "2WD",
        "decodeSource": source,
        "decodeMethod": "QXB",
        "confidence": source,
    }


def _map_vin_cache_entry(vin: str, entry: dict[str, Any]) -> dict[str, Any] | None:
    raw = entry.get("rawResponse") if isinstance(entry, dict) else None
    if isinstance(raw, dict):
        return _map_qxb_raw_to_decode(vin, raw)
    return _record_to_vin_decode({**entry, "vin": vin}, source="vin_cache_legacy")


def _record_to_vin_decode(rec: dict[str, Any], *, source: str) -> dict[str, Any] | None:
    vin = normalize_vin_strict(str(rec.get("vin") or ""))
    if not vin or not VIN_STRICT.match(vin):
        return None
    engine = str(rec.get("engineCode") or "").strip()
    trans = str(rec.get("transmissionCode") or "").strip()
    if not engine and not trans and not rec.get("model"):
        return None
    return {
        "ok": True,
        "vin": vin,
        "brand": rec.get("brand") or "",
        "brandSlug": rec.get("brandSlug") or "",
        "model": rec.get("model") or "",
        "year": rec.get("year") or None,
        "engineCode": engine,
        "transmissionCode": trans,
        "drivetrain": rec.get("drivetrain") or "2WD",
        "decodeSource": source,
        "decodeMethod": rec.get("decodeMethod") or "Cached",
        "confidence": rec.get("decodeConfidence") or source,
    }


def _lookup_cached_vin_decode(vin: str, *, root: Path | None = None) -> dict[str, Any] | None:
    """Local cache first — same VIN should never hit paid decode twice."""
    vin_norm = normalize_vin_strict(vin)
    if not vin_norm or not VIN_STRICT.match(vin_norm):
        return None

    mem = _VIN_DECODE_MEMORY.get(vin_norm)
    if mem:
        return dict(mem)

    cache_path = _vin_decode_cache_path(root)
    if cache_path.is_file():
        try:
            store = load_json(cache_path, {})
            hit = store.get(vin_norm)
            if isinstance(hit, dict) and hit.get("ok"):
                _VIN_DECODE_MEMORY[vin_norm] = dict(hit)
                return dict(hit)
        except Exception:
            pass

    facts_path = _approved_vin_facts_path(root)
    if facts_path.is_file():
        try:
            facts = load_json(facts_path, {})
            fact = facts.get(vin_norm)
            if isinstance(fact, dict):
                mapped = _record_to_vin_decode(fact, source="approved_vin_facts")
                if mapped:
                    _VIN_DECODE_MEMORY[vin_norm] = mapped
                    return mapped
        except Exception:
            pass

    approved_out = (root or ROOT) / "reports/qxb-approved-import.json"
    if approved_out.is_file():
        try:
            for rec in load_json(approved_out, []):
                if normalize_vin_strict(str(rec.get("vin") or "")) != vin_norm:
                    continue
                mapped = _record_to_vin_decode(rec, source="approved_import")
                if mapped and mapped.get("engineCode"):
                    _VIN_DECODE_MEMORY[vin_norm] = mapped
                    return mapped
        except Exception:
            pass

    legacy_cache = (root or ROOT) / "data/knowledge-base/vin-cache.json"
    if legacy_cache.is_file():
        try:
            entry = load_json(legacy_cache, {}).get(vin_norm)
            if isinstance(entry, dict):
                mapped = _map_vin_cache_entry(vin_norm, entry)
                if mapped and (mapped.get("engineCode") or mapped.get("model")):
                    _VIN_DECODE_MEMORY[vin_norm] = mapped
                    return mapped
        except Exception:
            pass

    return None


def _persist_vin_decode_cache(vin: str, body: dict[str, Any], *, root: Path | None = None) -> None:
    if not body.get("ok"):
        return
    if _vin_decode_cache_tier(body) <= 0:
        return
    vin_norm = normalize_vin_strict(vin)
    if not vin_norm:
        return
    payload = dict(body)
    payload["vin"] = vin_norm
    payload.setdefault("cachedAt", _iso_now())
    cache_path = _vin_decode_cache_path(root)
    store = load_json(cache_path, {}) if cache_path.is_file() else {}
    existing = store.get(vin_norm)
    if isinstance(existing, dict) and existing.get("ok"):
        payload = _merge_vin_decode_payload(existing, payload)
    _VIN_DECODE_MEMORY[vin_norm] = payload
    store[vin_norm] = payload
    save_json_atomic(cache_path, store)


def warm_vin_decode_cache_from_approved(*, root: Path | None = None) -> int:
    """Seed vin-decode-cache.json from prior approved import rows (no API calls)."""
    approved_out = (root or ROOT) / "reports/qxb-approved-import.json"
    if not approved_out.is_file():
        return 0
    cache_path = _vin_decode_cache_path(root)
    store = load_json(cache_path, {}) if cache_path.is_file() else {}
    added = 0
    for rec in load_json(approved_out, []):
        mapped = _record_to_vin_decode(rec, source="approved_import_seed")
        if not mapped or not mapped.get("engineCode"):
            continue
        vin = mapped["vin"]
        if vin in store and _vin_decode_cache_tier(store[vin]) >= 3:
            continue
        if vin in store and _vin_decode_cache_tier(mapped) <= _vin_decode_cache_tier(store[vin]):
            continue
        store[vin] = mapped
        _VIN_DECODE_MEMORY[vin] = mapped
        added += 1
    if added:
        save_json_atomic(cache_path, store)
    return added


def _vin_decode_retryable(body: dict[str, Any] | None, http_code: int = 0) -> bool:
    if http_code in VIN_DECODE_RETRY_HTTP:
        return True
    reason = str((body or {}).get("reason") or "").strip()
    if reason in VIN_DECODE_RETRY_REASONS:
        return True
    msg = str((body or {}).get("message") or (body or {}).get("error") or "").lower()
    return bool(re.search(r"\b(502|503|504|quota|限额|rate.?limit|too many)\b", msg))


def decode_vin_via_api(
    api_base: str,
    vin: str,
    *,
    max_retries: int | None = None,
    retry_backoff: float | None = None,
    root: Path | None = None,
) -> dict[str, Any]:
    """Decode VIN via asia-power.com — local cache first, then API (QXB is billed per call)."""
    vin_norm = str(vin or "").strip().upper()
    if len(vin_norm) != 17:
        return {"ok": False, "error": "VIN must be 17 characters", "vin": vin_norm}

    cached = _lookup_cached_vin_decode(vin_norm, root=root)
    if cached:
        return dict(cached)

    url = f"{api_base.rstrip('/')}/api/vin/decode"
    attempts = max(0, int(DEFAULT_VIN_DECODE_MAX_RETRIES if max_retries is None else max_retries)) + 1
    backoff = DEFAULT_VIN_DECODE_BACKOFF_SEC if retry_backoff is None else float(retry_backoff)
    last: dict[str, Any] = {"ok": False, "error": "decode failed", "vin": vin_norm}
    for attempt in range(attempts):
        try:
            req = urllib.request.Request(
                url,
                data=json.dumps({"vin": vin_norm}).encode(),
                headers={
                    "Content-Type": "application/json",
                    "User-Agent": "AsiaPower-QXB-Pipeline/1.5",
                },
                method="POST",
            )
            with urllib.request.urlopen(req, timeout=45) as res:
                body = json.loads(res.read().decode())
            if body.get("ok"):
                body.setdefault("vin", vin_norm)
                body.setdefault("decodeSource", "qxb_via_asia_power")
                _persist_vin_decode_cache(vin_norm, body, root=root)
                return body
            last = {
                "ok": False,
                "error": format_vin_decode_error(body),
                "reason": body.get("reason"),
                "vin": vin_norm,
            }
            if _vin_decode_retryable(body) and attempt < attempts - 1:
                time.sleep(backoff * (attempt + 1))
                continue
            return last
        except urllib.error.HTTPError as exc:
            try:
                detail = exc.read().decode(errors="replace")[:300]
            except Exception:
                detail = exc.reason or str(exc)
            last = {"ok": False, "error": detail or str(exc), "vin": vin_norm, "httpCode": exc.code}
            if _vin_decode_retryable(None, exc.code) and attempt < attempts - 1:
                time.sleep(backoff * (attempt + 1))
                continue
            return last
        except Exception as exc:
            last = {"ok": False, "error": str(exc), "vin": vin_norm}
            err_text = str(exc).lower()
            if attempt < attempts - 1 and re.search(r"\b(502|503|504|timed out|timeout)\b", err_text):
                time.sleep(backoff * (attempt + 1))
                continue
            return last
    return last


def apply_vin_decode_to_record(record: dict[str, Any], decode: dict[str, Any]) -> dict[str, Any]:
    """Merge decoded powertrain fields into a listing/submission record."""
    if not decode.get("ok"):
        return record
    for key in ("engineCode", "transmissionCode", "drivetrain", "fuelType", "gearboxModel", "displacement"):
        val = decode.get(key)
        if val not in (None, ""):
            record[key] = val
    if decode.get("brand") and not record.get("brand"):
        record["brand"] = decode["brand"]
    if decode.get("brandSlug") and not record.get("brandSlug"):
        record["brandSlug"] = decode["brandSlug"]
    decode_model = normalize_decode_model_name(
        str(decode.get("model") or ""),
        str(record.get("brand") or decode.get("brand") or ""),
        str((record.get("qxb") or {}).get("brandCn") or ""),
    )
    if decode_model:
        record["model"] = decode_model
    if decode.get("year") and not record.get("year"):
        record["year"] = decode["year"]
    record["decodeMethod"] = "Auto Decoded"
    if decode.get("confidence"):
        record["decodeConfidence"] = decode["confidence"]
    engine = record.get("engineCode") or ""
    if engine and record.get("brand") and record.get("model"):
        record["title"] = f"{record['brand']} {record['model']} {engine} Half Cut"
    return record


DISP_LITERS_RE = re.compile(r"(\d+\.\d+)\s*(?:L|升)", re.I)
DISP_TURBO_RE = re.compile(r"(\d+\.\d+)T", re.I)
DISP_NUM_RE = re.compile(r"(\d+\.\d+)\s*(?:L|升|MT|AT|CVT|手动|自动)", re.I)


def parse_displacement_liters(trim: str) -> float | None:
    for pat in (DISP_LITERS_RE, DISP_TURBO_RE, DISP_NUM_RE):
        m = pat.search(trim or "")
        if m:
            try:
                return float(m.group(1))
            except ValueError:
                continue
    return None


def infer_engine_from_trim(trim: str, brand: str, model: str) -> str:
    """Infer engine code from trim displacement when VIN decode + catalog miss."""
    disp = parse_displacement_liters(trim)
    b = brand.lower()
    m = model.lower()
    text = trim or ""
    if "mercedes" in b:
        lower = text.lower()
        compact = lower.replace(" ", "")
        if "ml" in lower or "m级" in text:
            if "350" in text:
                return "M272.967"
            if "320" in text:
                return "M272.964"
            if "500" in text:
                return "M113.964"
        if "c200" in compact or "c200k" in compact:
            return "M271.951"
        if "c180" in compact:
            return "M271.820"
        if "e200" in compact or "e200k" in compact:
            return "M271.860"
    if disp is None:
        return ""
    if b == "honda":
        if "fit" in m or "飞度" in text:
            if disp <= 1.4:
                return "L13Z"
            if disp <= 1.6:
                return "L15A7"
        if "city" in m or "锋范" in text:
            if disp <= 1.6:
                return "L15A1"
        if "accord" in m or "雅阁" in text:
            if disp >= 2.3:
                return "K24A8"
            if disp >= 1.9:
                return "R20A3"
        if "cr-v" in m or "crv" in m.replace("-", "") or "cr-v" in text.lower():
            if disp is not None and disp >= 2.3:
                return "K24Z4"
        if "odyssey" in m or "奥德赛" in text:
            if disp is not None and disp >= 2.3:
                return "K24W7"
        if "civic" in m or "思域" in text:
            if disp is not None and disp >= 1.7:
                return "R18A2"
        if "elysion" in m or "艾力绅" in text:
            if disp is not None and disp >= 2.3:
                return "K24Z5"
        if disp >= 1.9:
            return "R20A3"
    if b == "nissan":
        if "teana" in m or "天籁" in text:
            if disp is not None and disp >= 2.2:
                return "VQ35DE"
            if disp is not None and disp >= 1.9:
                return "VQ23DE"
            if re.search(r"\b230\b|2\.3", text):
                return "VQ23DE"
            if re.search(r"\b350\b|3\.5", text):
                return "VQ35DE"
        if "奇骏" in text or "x-trail" in m.replace("-", "").lower():
            if disp is not None and disp >= 2.3:
                return "QR25DE"
        if disp is not None and disp <= 1.7:
            return "HR16DE"
        if disp is not None and disp <= 2.1:
            return "MR20DE"
        if disp is not None and disp <= 2.5:
            return "QR25DE"
    if b == "hyundai":
        if "领翔" in text or "sonata" in m and "nf" in text.lower():
            if disp is not None and disp <= 2.1:
                return "G4KA"
        if "ix35" in m or "ix35" in text.lower():
            if disp is not None and disp >= 2.3:
                return "G4KC"
            if disp is not None and disp <= 2.1:
                return "G4NA"
        if "悦动" in text or "elantra" in m:
            if disp is not None and disp <= 1.7:
                return "G4FC"
        if "名图" in text or "mistra" in m:
            if disp is not None and disp <= 2.1:
                return "G4NA"
        if "胜达" in text or "santa" in m:
            if disp is not None and disp <= 2.5:
                return "G4KE" if "经典" in text else "G4KC"
        if disp is not None and disp <= 1.7:
            return "G4FC"
        if disp is not None and disp <= 2.1:
            return "G4NA"
        if disp is not None and disp <= 2.5:
            return "G4KC"
    if b in {"volkswagen", "大众"} or "大众" in text:
        lower = text.lower()
        if "passat" in m or "领驭" in text or "帕萨特" in text:
            if "1.8t" in lower or (disp is not None and disp <= 2.0):
                return "CEA"
        if "santana" in m or "桑塔纳" in m or "桑塔纳" in text or "志俊" in m or "志俊" in text:
            if disp is not None and disp <= 2.0:
                return "BWH"
        if "lavida" in m or "朗逸" in text:
            if disp is not None and disp <= 1.7:
                return "CPD"
    if b == "toyota":
        if "previa" in m or "普瑞维亚" in text:
            if disp is not None and disp >= 2.2:
                return "2AZ-FE"
        if "highlander" in m or "汉兰达" in text:
            if disp is not None and disp <= 2.8:
                return "2AR-FE" if disp >= 2.4 else "1AR-FE"
        if "rav4" in m or "荣放" in text:
            if disp is not None and disp >= 2.3:
                return "2AZ-FE"
        if "camry" in m or "凯美瑞" in text:
            if disp is not None and disp >= 2.3:
                return "2AZ-FE" if disp <= 2.5 else "2GR-FE"
            if disp is not None and disp >= 2.0:
                return "6AR-FSE"
        if "corolla" in m or "卡罗拉" in text:
            if "双擎" in text or "hybrid" in text.lower():
                return "2ZR-FXE"
            if disp is not None and disp <= 1.7:
                return "1ZR-FE"
        if "crown" in m or "皇冠" in text:
            if "2.0t" in text.lower() or (disp is not None and disp <= 2.1):
                return "8AR-FTS"
        if "vios" in m or "威驰" in text:
            if disp is not None and disp <= 1.4:
                return "1NR-FE"
            if disp is not None and disp <= 1.7:
                return "1NZ-FE"
        if "reiz" in m or "锐志" in text:
            if disp is not None and disp >= 2.4:
                return "4GR-FSE"
        if disp is not None and disp <= 1.7:
            return "1ZR-FE"
        if disp is not None and disp <= 2.1:
            return "1AZ-FE"
        if disp is not None and disp <= 2.5:
            return "2AZ-FE"
    if b == "mitsubishi":
        if "pajero" in m or "帕杰罗" in text:
            if disp is not None and disp <= 3.2:
                return "6G72"
        if "asx" in m or "劲炫" in text:
            if disp is not None and disp <= 2.1:
                return "4B11"
    if b == "suzuki":
        if "swift" in m or "雨燕" in text:
            if disp is not None and disp <= 1.4:
                return "M13A"
        if "alto" in m or "奥拓" in text:
            return "K10B"
        if "sx4" in m or "天语" in text:
            if disp is not None and disp <= 1.7:
                return "M16A"
    if "mercedes" in b:
        lower = text.lower()
        compact = lower.replace(" ", "")
        if "gl450" in compact or "gl 450" in lower:
            return "M273.923"
        if "glk300" in compact or "glk 300" in lower:
            return "M272.948"
        if "b200" in compact or "b 200" in lower:
            return "M266.940"
    if b == "kia":
        if "sportage" in m or "智跑" in text:
            if disp is not None and disp <= 2.1:
                return "G4KD"
        if "k5" in m or "k5" in text.lower():
            if disp is not None and disp <= 2.1:
                return "G4NA"
        if "赛拉图" in text or "cerato" in m:
            if disp <= 1.7:
                return "G4ED"
        if disp <= 1.7:
            return "G4FC"
        if disp is not None and disp <= 2.1:
            return "G4NA"
    return ""


def infer_default_transmission(trim: str, brand: str, model: str) -> str:
    """When trim omits 自动/手动, infer likely AT from brand/model era."""
    text = trim or ""
    year_m = re.search(r"(\d{4})", text)
    year = int(year_m.group(1)) if year_m else 0
    b = brand.lower()
    m = model.lower()
    if b == "toyota" and ("previa" in m or "普瑞维亚" in text):
        return "4AT" if year and year < 2010 else "6AT"
    if b == "honda" and ("accord" in m or "雅阁" in text):
        return "5AT" if year and year < 2013 else "CVT"
    if "mercedes" in b:
        compact = text.lower().replace(" ", "")
        if ("c200" in compact or "c180" in compact) and year and year < 2010:
            return "5AT"
        if "4matic" in text.lower() or "四驱" in text:
            return "7AT"
        return "7AT" if year and year >= 2005 else "5AT"
    if b == "hyundai":
        if "胜达" in text or "santa" in m:
            return "6MT" if "手动" in text else "6AT"
        if "四驱" in text or "4wd" in text.lower():
            return "6MT" if "手动" in text else "6AT"
    if b == "nissan" and ("奇骏" in text or "x-trail" in m.lower()):
        if "无级" in text or "cvt" in text.upper():
            return "CVT"
        return "6AT"
    return ""


def pick_fallback_engine(trim: str, brand: str, model: str, refs: list[dict[str, Any]]) -> str:
    """Use catalog refs only when model matches; else trim displacement."""
    matched = [r for r in refs if _model_match(str(r.get("model") or ""), model)]
    if matched:
        return str(matched[0].get("engineCode") or "").strip()
    return infer_engine_from_trim(trim, brand, model)


def infer_transmission_from_trim(trim: str) -> str:
    """Infer transmission code from Chinese trim text (e.g. 自动 → 6AT)."""
    text = trim or ""
    if "无级" in text or "CVT" in text.upper():
        return "CVT"
    if "双离合" in text or "DCT" in text.upper():
        return "DCT"
    if "手动" in text or re.search(r"\bMT\b", text, re.I):
        return "6MT" if re.search(r"6速|6MT|6挡", text, re.I) else "5MT"
    if "自动" in text or re.search(r"\bAT\b", text, re.I):
        if "7速" in text or "7AT" in text.upper():
            return "7AT"
        if "8速" in text:
            return "8AT"
        if "4速" in text or "4AT" in text.upper():
            return "4AT"
        return "6AT"
    return ""


def fallback_vin_decode(ctx: PipelineContext, fields: dict[str, Any]) -> dict[str, Any] | None:
    """When /api/vin/decode fails, infer powertrain from trim + similar catalog listings."""
    vin = str((fields.get("vin_info") or {}).get("vin") or "").strip().upper()
    if len(vin) != 17:
        return None
    trim = str((fields.get("source") or {}).get("trim") or "")
    brand = fields["brand"]
    model = fields["model"]

    def _catalog_engine() -> str:
        try:
            price_meta = estimate_half_cut_price_usd(
                brand=brand,
                brand_slug=fields["brand_slug"],
                model=model,
                year=(fields.get("source") or {}).get("year"),
                engine_code="",
                api_base=ctx.paths.api_base,
            )
            refs = price_meta.get("references") or []
            return pick_fallback_engine(trim, brand, model, refs)
        except Exception:
            return ""

    trans = infer_transmission_from_trim(trim) or infer_default_transmission(trim, brand, model)
    engine = infer_engine_from_trim(trim, brand, model) or _catalog_engine()

    if engine and not trans:
        trans = infer_default_transmission(trim, brand, model)
    if trans and not engine:
        engine = infer_engine_from_trim(trim, brand, model) or _catalog_engine()

    if not engine or not trans:
        return None
    return {
        "ok": True,
        "vin": vin,
        "engineCode": engine,
        "transmissionCode": trans,
        "drivetrain": "4WD" if ("四驱" in trim or "4wd" in trim.lower() or "4matic" in trim.lower()) else "2WD",
        "confidence": "trim_catalog_fallback",
        "decodeMethod": "Trim+Catalog",
    }


def merge_nameplate_powertrain(
    ctx: PipelineContext,
    fields: dict[str, Any],
    vin_info: dict | None,
    vin_decode: dict[str, Any] | None,
) -> dict[str, Any]:
    """铭牌发动机/变速箱优先于 VIN decode API 与 trim 猜测。"""
    merged: dict[str, Any] = dict(vin_decode or {})
    plate_engine = str((vin_info or {}).get("engine_code") or "").strip().upper()
    plate_trans = str((vin_info or {}).get("transmission_code") or "").strip().upper()

    plate_path = str((vin_info or {}).get("image_path") or "")
    if plate_path and Path(plate_path).is_file() and not plate_engine:
        try:
            from importlib.machinery import SourceFileLoader

            ocr = SourceFileLoader("ocr", str(ROOT / "scripts/ocr-qxb-vins.py")).load_module()
            ocr_pt = ocr.nameplate_powertrain_for(Path(plate_path))
            plate_engine = str(ocr_pt.get("engineCode") or "").strip().upper()
            if not plate_trans:
                plate_trans = str(ocr_pt.get("transmissionCode") or "").strip().upper()
        except Exception:
            pass

    if plate_engine:
        merged["engineCode"] = plate_engine
        merged["decodeMethod"] = "Nameplate"
        merged["ok"] = True
        merged["confidence"] = "nameplate"

    if plate_trans:
        merged["transmissionCode"] = plate_trans

    if merged.get("engineCode") and not merged.get("transmissionCode"):
        trim = str((fields.get("source") or {}).get("trim") or "")
        trans = infer_transmission_from_trim(trim) or infer_default_transmission(
            trim, fields["brand"], fields["model"],
        )
        if trans:
            merged["transmissionCode"] = trans
            merged.setdefault("drivetrain", "2WD")

    if merged.get("engineCode") and merged.get("transmissionCode"):
        merged["ok"] = True

    return merged


def powertrain_blockers(vin: str | None, vin_decode: dict[str, Any] | None) -> list[str]:
    """QXB rule: engine + transmission required; nameplate engine overrides decode."""
    if not vin:
        return ["no VIN — cannot derive engine/transmission (Manual Entry not allowed for QXB live upload)"]
    decode = vin_decode or {}
    if decode.get("decodeMethod") == "Nameplate" and str(decode.get("engineCode") or "").strip():
        if not str(decode.get("transmissionCode") or "").strip():
            return ["transmissionCode missing — infer from trim or enter manually"]
        return []
    if not decode.get("ok"):
        err = decode.get("error") or "VIN decode API unavailable"
        return [f"VIN decode failed for {vin}: {err}"]
    blockers: list[str] = []
    if not str(vin_decode.get("engineCode") or "").strip():
        blockers.append("engineCode missing after VIN decode")
    if not str(vin_decode.get("transmissionCode") or "").strip():
        blockers.append("transmissionCode missing after VIN decode")
    return blockers


def resolve_submission_id(row: int, queue_entry: dict[str, Any] | None = None) -> str:
    """Unique submission id; after /qxb reupload use -R{n} so rejected rows can resubmit."""
    count = int((queue_entry or {}).get("resubmitCount") or 0)
    base = f"QXB-{row:04d}"
    return f"{base}-R{count}" if count > 0 else base


def build_record(
    source: dict[str, Any],
    brand: str,
    brand_slug: str,
    model: str,
    model_key: str,
    photos: list[dict],
    vin_info: dict | None,
    *,
    vin_decode: dict[str, Any] | None = None,
    submission_id: str | None = None,
    price_meta: dict[str, Any] | None = None,
) -> dict[str, Any]:
    row = source["row"]
    vin = (vin_info or {}).get("vin", "")
    decode = vin_decode if vin_decode and vin_decode.get("ok") else None
    if decode:
        decode_brand = str(decode.get("brand") or "").strip()
        if decode_brand:
            brand = decode_brand
            brand_slug = slugify(decode_brand)
        decode_model = normalize_decode_model_name(
            str(decode.get("model") or ""),
            brand,
            str(source.get("brand_cn") or ""),
        )
        if decode_model:
            model = decode_model
            model_key = decode_model
    engine_code = (decode or {}).get("engineCode") or ""
    trans_code = (decode or {}).get("transmissionCode") or ""
    drivetrain = (decode or {}).get("drivetrain") or "2WD"
    decode_method = (
        "Nameplate"
        if decode and decode.get("decodeMethod") == "Nameplate"
        else ("Auto Decoded" if decode and engine_code else ("QXB OCR" if vin else "Manual Entry"))
    )
    notes = [
        "汽修宝批量导入。",
        "里程数为系统默认 99,999 km，仅为占位，不代表真实里程。",
        f"原始车型: {source['trim']}",
    ]
    if source.get("description"):
        notes.append(f"原始说明: {source['description']}")
    if (vin_info or {}).get("confidence"):
        notes.append(f"VIN OCR confidence: {vin_info['confidence']}")
    if decode and decode.get("decodeMethod") == "Nameplate":
        notes.append(f"Nameplate engine: {decode.get('engineCode')}")
    if decode and engine_code and decode.get("decodeMethod") != "Nameplate":
        notes.append(f"VIN decode: engine {engine_code}, transmission {trans_code or '—'}")
    if price_meta and price_meta.get("note"):
        notes.append(str(price_meta["note"]))
    if not decode and vin:
        notes.append("WARN: VIN decode did not return powertrain — do not submit until enriched")
    stock_id = f"QXB{row:04d}"
    sid = submission_id or f"QXB-{row:04d}"
    price_usd = (price_meta or {}).get("priceUsd")
    title = f"{brand} {model} {engine_code} Half Cut".strip() if engine_code else f"{brand} {model} Half Cut"
    slug = "-".join(filter(bool, [
        brand_slug, slugify(model), str(source.get("year") or ""), "half-cut", stock_id.lower(),
    ]))
    return {
        "stockId": stock_id,
        "vin": vin,
        "decodeMethod": decode_method,
        "decodeConfidence": (vin_info or {}).get("confidence") or (decode or {}).get("confidence"),
        "vehicleCondition": "Half Cut",
        "vehicleCategory": "passenger",
        "truckPartType": "",
        "machineryType": "",
        "brand": brand,
        "brandSlug": brand_slug,
        "model": model,
        "year": source.get("year") or (decode or {}).get("year") or None,
        "engineCode": engine_code,
        "transmissionCode": trans_code,
        "drivetrain": drivetrain,
        "mileage": "99,999 km",
        "priceUsd": price_usd,
        "priceEstimated": bool((price_meta or {}).get("priceEstimated")),
        "origin": "China",
        "status": "Available",
        "title": title,
        "slug": slug,
        "photos": photos,
        "video": None,
        "videoUrl": "",
        "includedParts": ["Engine & gearbox assembly", "Front clip"],
        "shortDescription": f"{source.get('year') or ''} {brand} {model} — QXB sourced listing via AsiaPower.".strip(),
        "supplierVerified": True,
        "supplierName": "汽修宝",
        "supplierPhone": "16638801930",
        "supplierCity": "",
        "submissionId": sid,
        "approvedAt": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "notes": "\n".join(notes),
        "qxb": {
            "row": row,
            "brandCn": source["brand_cn"],
            "modelKey": model_key,
            "description": source.get("description") or "",
            "priceEstimate": price_meta or None,
        },
    }


@dataclass
class PipelineContext:
    paths: QxbPaths
    sources: dict[int, dict[str, Any]]
    manifest: dict[int, list[dict[str, str]]]
    vins: dict[int, dict[str, str]]
    cn_to_en: dict[str, str]
    cn_to_slug: dict[str, str]
    upload_state: dict[str, Any]
    queue: dict[str, Any]
    approved: list[dict[str, Any]]


def load_context(paths: QxbPaths | None = None) -> PipelineContext:
    p = paths or QxbPaths()
    sources = {s["row"]: s for s in load_rows(p.xlsx)}
    cn_to_en, cn_to_slug = load_brand_maps(p.brand_dict)
    approved = load_json(p.approved_out, [])
    if not isinstance(approved, list):
        approved = []
    queue = load_json(p.agent_queue, {"version": "1.0", "updated_at": _now(), "rows": {}})
    if "rows" not in queue:
        queue["rows"] = {}
    return PipelineContext(
        paths=p,
        sources=sources,
        manifest=load_manifest(p.manifest),
        vins=load_vins(p.vin_csv),
        cn_to_en=cn_to_en,
        cn_to_slug=cn_to_slug,
        upload_state=load_json(p.upload_state, {"uploads": {}}),
        queue=queue,
        approved=approved,
    )


def _queue_row(ctx: PipelineContext, row: int) -> dict[str, Any]:
    key = str(row)
    entry = ctx.queue["rows"].setdefault(key, {
        "row": row,
        "status": "pending",
        "issues": [],
        "history": [],
    })
    return entry


def _append_history(entry: dict[str, Any], action: str, detail: str = "") -> None:
    entry.setdefault("history", []).append({
        "at": _iso_now(),
        "action": action,
        "detail": detail[:500],
    })


def save_queue(ctx: PipelineContext) -> None:
    ctx.queue["updated_at"] = _now()
    save_json_atomic(ctx.paths.agent_queue, ctx.queue)


def park_row(
    ctx: PipelineContext,
    row: int,
    *,
    category: str,
    note: str = "",
    tier: str = "later",
) -> dict[str, Any]:
    """Mark row as parked — skip in auto batch until unparked."""
    if category not in PARK_CATEGORIES:
        raise ValueError(f"unknown park category: {category}")
    entry = _queue_row(ctx, row)
    prev_status = entry.get("status")
    entry["status"] = "parked"
    entry["park"] = {
        "category": category,
        "label": PARK_CATEGORIES[category],
        "tier": tier,
        "note": note,
        "markedAt": _iso_now(),
        "previousStatus": prev_status,
    }
    entry["updated_at"] = _iso_now()
    _append_history(entry, "parked", f"{category}: {note or PARK_CATEGORIES[category]}")
    save_queue(ctx)
    return entry


def unpark_row(ctx: PipelineContext, row: int) -> dict[str, Any]:
    """Restore row from parked to previous status (default pending)."""
    entry = _queue_row(ctx, row)
    prev = (entry.get("park") or {}).get("previousStatus") or "pending"
    entry.pop("park", None)
    entry["status"] = prev if prev != "parked" else "pending"
    entry["updated_at"] = _iso_now()
    _append_history(entry, "unparked", f"restored → {entry['status']}")
    save_queue(ctx)
    return entry


def list_parked_rows(ctx: PipelineContext) -> list[dict[str, Any]]:
    rows = ctx.queue.get("rows") or {}
    out: list[dict[str, Any]] = []
    for key in sorted(rows, key=lambda k: int(k)):
        entry = rows[key]
        if entry.get("status") != "parked":
            continue
        park = entry.get("park") or {}
        src = ctx.sources.get(int(key), {})
        out.append({
            "row": int(key),
            "stockId": f"QXB{int(key):04d}",
            "category": park.get("category"),
            "label": park.get("label"),
            "tier": park.get("tier", "later"),
            "note": park.get("note", ""),
            "trim": src.get("trim", ""),
            "markedAt": park.get("markedAt"),
        })
    return out


def format_parked_report(ctx: PipelineContext) -> str:
    parked = list_parked_rows(ctx)
    if not parked:
        return "无暂缓行 — 全部可继续处理。"
    easy = [p for p in parked if p.get("tier") == "easy"]
    later = [p for p in parked if p.get("tier") != "easy"]
    lines = [
        f"QXB 暂缓标记 ({len(parked)} 行) — 先处理容易的，这些稍后再来",
        "",
    ]
    if easy:
        lines.append("【easy — 补推/小修即可】")
        for p in easy:
            lines.append(
                f"  QXB{p['row']:04d} [{p.get('category')}] {p.get('label')} — {p.get('note') or p.get('trim','')[:40]}"
            )
        lines.append("")
    if later:
        lines.append("【later — 需 VIN/相册/CEO 操作】")
        for p in later:
            lines.append(
                f"  QXB{p['row']:04d} [{p.get('category')}] {p.get('label')} — {p.get('note') or p.get('trim','')[:40]}"
            )
    lines.append("")
    lines.append("解除暂缓: /qxb unpark <row>")
    return "\n".join(lines)


def set_row_status(
    ctx: PipelineContext,
    row: int,
    status: str,
    *,
    issue: str | None = None,
    clear_issues: bool = False,
) -> None:
    if status not in ROW_STATUSES:
        raise ValueError(f"invalid status: {status}")
    entry = _queue_row(ctx, row)
    entry["status"] = status
    entry["updated_at"] = _iso_now()
    if clear_issues:
        entry["issues"] = []
    if issue:
        entry.setdefault("issues", []).append({"at": _iso_now(), "text": issue})
    _append_history(entry, f"status:{status}", issue or "")
    save_queue(ctx)


def resolve_source(ctx: PipelineContext, row: int) -> dict[str, Any] | None:
    return ctx.sources.get(row)


def resolve_listing_fields(ctx: PipelineContext, row: int) -> dict[str, Any] | None:
    source = resolve_source(ctx, row)
    if not source:
        return None
    brand = ctx.cn_to_en.get(source["brand_cn"], source["brand_cn"])
    brand_slug = ctx.cn_to_slug.get(source["brand_cn"], slugify(brand))
    model_key, latin = parse_model(source["brand_cn"], source["trim"])
    model = model_english(ctx.paths.model_dict, brand_slug, model_key, latin)
    photos = ctx.manifest.get(row, [])
    if bulk_upload_mode():
        picks, photo_pick_meta = all_manifest_photo_picks(photos)
    else:
        picks, photo_pick_meta = selected_photos(photos, ctx.vins.get(row), row=row)
    vin_info = ctx.vins.get(row)
    fields: dict[str, Any] = {
        "source": source,
        "brand": brand,
        "brand_slug": brand_slug,
        "model": model,
        "model_key": model_key,
        "photo_count": len(photos),
        "picks": picks,
        "photo_pick_meta": photo_pick_meta,
        "vin_info": vin_info,
        "vin_decode": None,
    }
    if vin_info and vin_info.get("vin"):
        if batch_skip_vin_decode():
            vin_decode = fallback_vin_decode(ctx, fields) or {
                "ok": False,
                "error": "decode skipped (batch mode)",
                "vin": vin_info["vin"],
            }
        else:
            vin_decode = decode_vin_via_api(ctx.paths.api_base, vin_info["vin"], root=ctx.paths.root)
            partial_decode = bool(
                vin_decode.get("ok")
                and (
                    not str(vin_decode.get("engineCode") or "").strip()
                    or not str(vin_decode.get("transmissionCode") or "").strip()
                )
            )
            if partial_decode:
                pre_fields = dict(fields)
                pre_fields["vin_decode"] = vin_decode
                apply_decode_listing_identity(pre_fields)
                for key in ("brand", "brand_slug", "model", "model_key"):
                    fields[key] = pre_fields[key]
            if not (
                vin_decode.get("ok")
                and vin_decode.get("engineCode")
                and vin_decode.get("transmissionCode")
            ):
                if (
                    partial_decode
                    and str(vin_decode.get("transmissionCode") or "").strip()
                    and not str(vin_decode.get("engineCode") or "").strip()
                ):
                    inferred_engine = infer_engine_from_trim(
                        str(source.get("trim") or ""),
                        fields["brand"],
                        fields["model"],
                    )
                    if inferred_engine:
                        vin_decode = dict(vin_decode)
                        vin_decode["engineCode"] = inferred_engine
                if not (
                    vin_decode.get("ok")
                    and str(vin_decode.get("engineCode") or "").strip()
                    and str(vin_decode.get("transmissionCode") or "").strip()
                ):
                    fb = fallback_vin_decode(ctx, fields)
                    if fb:
                        vin_decode = fb
                    elif partial_decode and str(vin_decode.get("engineCode") or "").strip():
                        vin_decode = dict(vin_decode)
                    elif partial_decode:
                        inferred_engine = infer_engine_from_trim(
                            str(source.get("trim") or ""),
                            fields["brand"],
                            fields["model"],
                        )
                        if inferred_engine:
                            vin_decode = dict(vin_decode)
                            vin_decode["engineCode"] = inferred_engine
                            if not str(vin_decode.get("transmissionCode") or "").strip():
                                vin_decode["transmissionCode"] = (
                                    infer_transmission_from_trim(source.get("trim") or "")
                                    or infer_default_transmission(
                                        source.get("trim") or "",
                                        fields["brand"],
                                        fields["model"],
                                    )
                                )
        fields["vin_decode"] = merge_nameplate_powertrain(ctx, fields, vin_info, vin_decode)
    apply_decode_listing_identity(fields)
    return fields


def parse_vin_ocr_confidence(raw: str | None) -> tuple[str, float | None]:
    """Map OCR confidence label (e.g. very_high:gray:rot0) to score 0–1."""
    text = str(raw or "").strip()
    if not text or text.lower() == "none":
        return "none", 0.0
    try:
        val = float(text)
        return text, max(0.0, min(1.0, val))
    except ValueError:
        pass
    lower = text.lower()
    if lower.startswith("very_high"):
        return text, 0.95
    if lower.startswith("high"):
        return text, 0.8
    if lower.startswith("medium") or lower.startswith("med"):
        return text, 0.6
    if lower.startswith("low"):
        return text, 0.3
    return text, None


def inspect_row(ctx: PipelineContext, row: int) -> dict[str, Any]:
    """Inspect one Excel row — blockers, warnings, and preview fields."""
    fields = resolve_listing_fields(ctx, row)
    entry = _queue_row(ctx, row)
    blockers: list[str] = []
    warnings: list[str] = []

    if not fields:
        blockers.append(f"Excel row {row} not found in {ctx.paths.xlsx}")
    else:
        source = fields["source"]
        min_photos = 1 if bulk_upload_mode() else 3
        if fields["photo_count"] < min_photos:
            blockers.append(f"only {fields['photo_count']} local photos (need ≥{min_photos})")
        if not bulk_upload_mode() and len(fields["picks"]) < 3:
            blockers.append(f"only {len(fields['picks'])} usable photo slots (need ≥3)")
        elif bulk_upload_mode() and len(fields["picks"]) < 1:
            blockers.append("no usable local photo files")
        if source["brand_cn"] not in ctx.cn_to_en:
            warnings.append(f"brand «{source['brand_cn']}» not in brand-dictionary — using raw name")
        if not fields["vin_info"]:
            warnings.append("no VIN in OCR results — will use Manual Entry")
        else:
            conf_label, conf_score = parse_vin_ocr_confidence(fields["vin_info"].get("confidence"))
            if conf_score is not None and conf_score < 0.5:
                warnings.append(f"low VIN OCR confidence: {conf_label}")
            vin_raw = fields["vin_info"].get("vin") or ""
            if not VIN_STRICT.match(vin_raw):
                blockers.append(
                    f"VIN invalid for admin approval: {vin_raw} "
                    "(I/O/Q not allowed — likely OCR 0↔O misread)"
                )
        pick_meta = fields.get("photo_pick_meta") or {}
        if not bulk_upload_mode() and pick_meta.get("method") == "heuristic_v1":
            warnings.append(
                f"photo slots via vision heuristics (confidence={pick_meta.get('confidence')}) — "
                "use /qxb preview to verify before live upload"
            )
        if not bulk_upload_mode() and pick_meta.get("confidence") == "low":
            warnings.append("photo slot confidence LOW — CEO preview required before upload")
        vin = (fields["vin_info"] or {}).get("vin") if fields else None
        vin_decode = fields.get("vin_decode") if fields else None
        blockers.extend(powertrain_blockers(vin, vin_decode))
        if vin:
            dup = find_duplicate_vin(ctx, vin, exclude_row=row)
            if dup:
                blockers.append(format_duplicate_vin_message(dup))
        if vin_decode and vin_decode.get("ok") and not batch_skip_vin_decode():
            pm = estimate_half_cut_price_usd(
                brand=fields["brand"],
                brand_slug=fields["brand_slug"],
                model=fields["model"],
                year=fields["source"].get("year"),
                engine_code=vin_decode.get("engineCode") or "",
                api_base=ctx.paths.api_base,
            )
            fields["price_meta"] = pm

    dup_info = None
    if fields:
        vin_for_dup = (fields.get("vin_info") or {}).get("vin")
        if vin_for_dup:
            dup_info = find_duplicate_vin(ctx, vin_for_dup, exclude_row=row)

    ready = not blockers
    result = {
        "row": row,
        "stockId": f"QXB{row:04d}" if fields else None,
        "status": entry.get("status", "pending"),
        "ready": ready,
        "blockers": blockers,
        "warnings": warnings,
        "issues": entry.get("issues") or [],
        "duplicateVin": dup_info,
        "duplicateVinMessage": format_duplicate_vin_message(dup_info) if dup_info else "",
        "fields": {
            "brand": fields["brand"] if fields else None,
            "model": fields["model"] if fields else None,
            "trim": fields["source"]["trim"] if fields else None,
            "photoCount": fields["photo_count"] if fields else 0,
            "photoSlots": [{"label": p["label"], "path": p["path"]} for p in (fields["picks"] if fields else [])],
            "vin": (fields["vin_info"] or {}).get("vin") if fields else None,
            "vinConfidence": (
                parse_vin_ocr_confidence((fields["vin_info"] or {}).get("confidence"))[0]
                if fields and fields.get("vin_info") else None
            ),
            "engineCode": (fields.get("vin_decode") or {}).get("engineCode") if fields else None,
            "transmissionCode": (fields.get("vin_decode") or {}).get("transmissionCode") if fields else None,
            "drivetrain": (fields.get("vin_decode") or {}).get("drivetrain") if fields else None,
            "decodeMethod": (fields.get("vin_decode") or {}).get("decodeMethod")
            if fields and (fields.get("vin_decode") or {}).get("ok") else None,
            "priceUsd": (fields.get("price_meta") or {}).get("priceUsd") if fields else None,
            "priceEstimated": (fields.get("price_meta") or {}).get("priceEstimated") if fields else None,
        },
        "sources": {
            "xlsx": str(ctx.paths.xlsx),
            "manifest": str(ctx.paths.manifest),
            "vinCsv": str(ctx.paths.vin_csv),
        },
    }
    if ready:
        set_row_status(ctx, row, "inspected")
    _append_history(entry, "inspect", "; ".join(blockers + warnings) or "ok")
    save_queue(ctx)
    return result


def _listing_build_extras(
    ctx: PipelineContext,
    row: int,
    fields: dict[str, Any],
) -> dict[str, Any]:
    entry = (ctx.queue.get("rows") or {}).get(str(row)) or {}
    decode = fields.get("vin_decode") or {}
    price_meta = estimate_half_cut_price_usd(
        brand=fields["brand"],
        brand_slug=fields["brand_slug"],
        model=fields["model"],
        year=fields["source"].get("year"),
        engine_code=(decode or {}).get("engineCode") or "",
        api_base=ctx.paths.api_base,
    )
    return {
        "submission_id": resolve_submission_id(row, entry),
        "price_meta": price_meta,
        "vin_decode": fields.get("vin_decode"),
    }


def prepare_row(ctx: PipelineContext, row: int, *, max_photos: int | None = None) -> dict[str, Any]:
    """Dry-run listing JSON (local photo paths, no R2 upload)."""
    inspection = inspect_row(ctx, row)
    if not inspection["ready"]:
        return {"ok": False, "inspection": inspection}

    fields = resolve_listing_fields(ctx, row)
    assert fields is not None
    photo_cap = max_photos if max_photos is not None else DEFAULT_UPLOAD_MAX_PHOTOS
    cap = _effective_max_photos(fields, photo_cap)
    preview_photos = [
        {
            "label": p["label"],
            "url": f"file://{Path(p['path']).resolve()}",
            "fileName": Path(p["path"]).name,
            "mimeType": mimetypes.guess_type(p["path"])[0] or "image/jpeg",
            "size": Path(p["path"]).stat().st_size,
        }
        for p in fields["picks"][:cap]
    ]
    extras = _listing_build_extras(ctx, row, fields)
    record = build_record(
        fields["source"],
        fields["brand"],
        fields["brand_slug"],
        fields["model"],
        fields["model_key"],
        preview_photos,
        fields["vin_info"],
        **extras,
    )
    set_row_status(ctx, row, "prepared")
    return {"ok": True, "dryRun": True, "record": record, "inspection": inspection}


def _http_error_detail(exc: urllib.error.HTTPError) -> str:
    try:
        body = exc.read().decode(errors="replace")[:300]
        parsed = json.loads(body) if body.strip().startswith("{") else None
        if isinstance(parsed, dict) and parsed.get("error"):
            return str(parsed["error"])
        return body or exc.reason
    except Exception:
        return exc.reason or str(exc)


UPLOAD_RETRY_HTTP_CODES = frozenset({423, 429, 503})
DEFAULT_UPLOAD_MAX_RETRIES = int(os.environ.get("QXB_UPLOAD_RETRY_MAX", "3"))
DEFAULT_UPLOAD_RETRY_BACKOFF_SEC = float(os.environ.get("QXB_UPLOAD_RETRY_BACKOFF", "10"))
# CEO 审核页：限流时立即返回，不长时间阻塞
REVIEW_UI_UPLOAD_MAX_RETRIES = int(os.environ.get("QXB_REVIEW_UPLOAD_RETRY_MAX", "0"))
REVIEW_UI_UPLOAD_RETRY_BACKOFF_SEC = float(os.environ.get("QXB_REVIEW_UPLOAD_RETRY_BACKOFF", "0"))


def _upload_retry_opts(
    max_retries: int | None = None,
    retry_backoff: float | None = None,
) -> dict[str, Any]:
    return {
        "max_retries": DEFAULT_UPLOAD_MAX_RETRIES if max_retries is None else int(max_retries),
        "retry_backoff": (
            DEFAULT_UPLOAD_RETRY_BACKOFF_SEC if retry_backoff is None else float(retry_backoff)
        ),
    }


def is_upload_rate_limited(message: str | None) -> bool:
    """True when error text indicates upload-token / presign throttling."""
    if not message:
        return False
    if re.search(r"\bHTTP (423|429|503)\b", message):
        return True
    return bool(re.search(r"\b(Too many|Locked)\b", message, re.I))


def _upload_rate_limit_hint(code: int, *, exhausted: bool) -> str:
    if code not in UPLOAD_RETRY_HTTP_CODES:
        return ""
    if exhausted:
        return (
            f"\n\n上传接口限流 (HTTP {code})。"
            f"\n请等待约 30 秒–1 分钟后重试。"
        )
    return f"\n\nHTTP {code} — 上传限流，正在自动重试…"


def _request_step_name(url: str) -> str:
    if "upload-token" in url:
        return "upload-token"
    if "presign" in url:
        return "upload/presign"
    return url


def _request_json(
    url: str,
    method: str,
    headers: dict,
    body: dict | None = None,
    *,
    max_retries: int = 0,
    retry_backoff: float = DEFAULT_UPLOAD_RETRY_BACKOFF_SEC,
) -> Any:
    data = None if body is None else json.dumps(body).encode()
    step = _request_step_name(url)
    attempts = max(0, int(max_retries)) + 1
    last_code = 0
    for attempt in range(attempts):
        req = urllib.request.Request(url, data=data, method=method, headers={
            "User-Agent": "AsiaPower-QXB-Agent/1.0",
            **headers,
        })
        if body is not None:
            req.add_header("Content-Type", "application/json")
        try:
            with urllib.request.urlopen(req, timeout=30) as res:
                return json.loads(res.read().decode())
        except urllib.error.HTTPError as exc:
            last_code = exc.code
            detail = _http_error_detail(exc)
            if exc.code in UPLOAD_RETRY_HTTP_CODES and attempt < attempts - 1:
                time.sleep(retry_backoff * (attempt + 1))
                continue
            hint = ""
            if exc.code == 403:
                hint = (
                    "\n\n403 Forbidden — 生产站拒绝了 X-Supplier-Key。"
                    "\n请确认本机 .env 的 SUPPLIER_UPLOAD_KEY 与 asia-power.com 服务器上的值完全一致。"
                    "\n（在服务器 .env 查看；不要用旧 key 或占位符。）"
                )
            elif exc.code == 401:
                hint = "\n\n401 — upload token 无效或过期，请重试 process。"
            elif exc.code in UPLOAD_RETRY_HTTP_CODES:
                hint = _upload_rate_limit_hint(exc.code, exhausted=True)
            raise RuntimeError(f"HTTP {exc.code} on {step}: {detail}{hint}") from exc
    raise RuntimeError(f"HTTP {last_code} on {step}: upload retries exhausted")


def check_upload_auth(paths: QxbPaths | None = None) -> dict[str, Any]:
    """Probe production upload-token endpoint (does not upload files)."""
    p = paths or QxbPaths()
    key = os.environ.get("SUPPLIER_UPLOAD_KEY", "")
    result: dict[str, Any] = {
        "apiBase": p.api_base,
        "keyConfigured": bool(key),
        "keyLength": len(key),
        "ok": False,
    }
    if not key:
        result["error"] = "SUPPLIER_UPLOAD_KEY not set in environment (.env)"
        return result
    try:
        token_resp = _request_json(
            f"{p.api_base}/api/half-cuts/upload-token",
            "POST",
            {"X-Supplier-Key": key},
            {},
        )
        result["ok"] = True
        result["tokenIssued"] = bool(token_resp.get("token"))
        return result
    except RuntimeError as exc:
        result["error"] = str(exc)
        return result


def format_upload_auth_report(data: dict[str, Any]) -> str:
    lines = [
        "QXB upload auth check",
        f"API: {data.get('apiBase')}",
        f"SUPPLIER_UPLOAD_KEY: {'set' if data.get('keyConfigured') else 'MISSING'}"
        + (f" (len={data.get('keyLength')})" if data.get("keyConfigured") else ""),
    ]
    if data.get("ok"):
        lines.append("Result: OK — production accepted supplier key, token issued.")
        lines.append("You may run: /qxb process <row> --live approved")
    else:
        lines.append(f"Result: FAILED\n{data.get('error', 'unknown')}")
    return "\n".join(lines)


def _should_use_server_upload(err: BaseException | str) -> bool:
    msg = str(err or "").lower()
    return (
        "direct upload not configured" in msg
        or "upload/presign" in msg
        or "presign failed" in msg
    )


def _upload_file_via_server(
    api_base: str,
    path: Path,
    label: str,
    supplier_key: str,
    token: str,
    *,
    max_retries: int | None = None,
    retry_backoff: float | None = None,
) -> dict[str, Any]:
    from urllib3.filepost import encode_multipart_formdata

    mime = mimetypes.guess_type(str(path))[0] or "image/jpeg"
    size = path.stat().st_size
    fields: list[tuple[Any, ...]] = []
    if label:
        fields.append(("label", label))
    fields.append(("file", (path.name, path.read_bytes(), mime)))
    body, content_type = encode_multipart_formdata(fields)
    url = f"{api_base}/api/half-cuts/upload/photo"
    attempts = max(0, int(max_retries if max_retries is not None else DEFAULT_UPLOAD_MAX_RETRIES)) + 1
    backoff = retry_backoff if retry_backoff is not None else DEFAULT_UPLOAD_RETRY_BACKOFF_SEC
    last_code = 0
    for attempt in range(attempts):
        req = urllib.request.Request(
            url,
            data=body,
            method="POST",
            headers={
                "Content-Type": content_type,
                "X-Supplier-Key": supplier_key,
                "X-Upload-Token": token,
                "User-Agent": "AsiaPower-QXB-Agent/1.0",
            },
        )
        try:
            with urllib.request.urlopen(req, timeout=120) as res:
                saved = json.loads(res.read().decode())
            return {
                "label": label,
                "url": saved.get("url") or saved.get("publicUrl") or "",
                "fileName": path.name,
                "mimeType": mime,
                "size": size,
            }
        except urllib.error.HTTPError as exc:
            last_code = exc.code
            detail = _http_error_detail(exc)
            if exc.code in UPLOAD_RETRY_HTTP_CODES and attempt < attempts - 1:
                time.sleep(backoff * (attempt + 1))
                continue
            hint = _upload_rate_limit_hint(exc.code, exhausted=True) if exc.code in UPLOAD_RETRY_HTTP_CODES else ""
            raise RuntimeError(f"HTTP {exc.code} on upload/photo: {detail}{hint}") from exc
    raise RuntimeError(f"HTTP {last_code} on upload/photo: upload retries exhausted")


def _upload_file(
    api_base: str,
    path: Path,
    label: str,
    supplier_key: str,
    token: str,
    *,
    max_retries: int | None = None,
    retry_backoff: float | None = None,
) -> dict[str, Any]:
    mime = mimetypes.guess_type(str(path))[0] or "image/jpeg"
    size = path.stat().st_size
    retry_opts = _upload_retry_opts(max_retries, retry_backoff)
    try:
        presign = _request_json(
            f"{api_base}/api/half-cuts/upload/presign",
            "POST",
            {"X-Supplier-Key": supplier_key, "X-Upload-Token": token},
            {"kind": "photo", "mimeType": mime, "filename": path.name, "size": size, "label": label},
            **retry_opts,
        )
        req = urllib.request.Request(
            presign["uploadUrl"],
            data=path.read_bytes(),
            method="PUT",
            headers={"Content-Type": presign.get("mimeType") or mime},
        )
        with urllib.request.urlopen(req, timeout=120) as res:
            if res.status >= 300:
                raise RuntimeError(f"R2 upload failed: {res.status}")
        return {
            "label": label,
            "url": presign["url"],
            "fileName": path.name,
            "mimeType": mime,
            "size": size,
        }
    except RuntimeError as exc:
        if not _should_use_server_upload(exc):
            raise
        return _upload_file_via_server(
            api_base,
            path,
            label,
            supplier_key,
            token,
            max_retries=max_retries,
            retry_backoff=retry_backoff,
        )


def _get_upload_token(
    api_base: str,
    supplier_key: str,
    *,
    max_retries: int | None = None,
    retry_backoff: float | None = None,
) -> str:
    return _request_json(
        f"{api_base}/api/half-cuts/upload-token",
        "POST",
        {"X-Supplier-Key": supplier_key},
        {},
        **_upload_retry_opts(max_retries, retry_backoff),
    )["token"]


def _merge_approved(ctx: PipelineContext, record: dict[str, Any]) -> None:
    stock_id = record.get("stockId")
    ctx.approved = [r for r in ctx.approved if r.get("stockId") != stock_id]
    ctx.approved.append(record)
    save_json_atomic(ctx.paths.approved_out, ctx.approved)


def prepare_row_reupload(ctx: PipelineContext, row: int) -> dict[str, Any]:
    """Reset a row for corrected photo re-upload (clears approved JSON + R2 cache keys)."""
    stock_id = f"QXB{row:04d}"
    before = len(ctx.approved)
    ctx.approved = [r for r in ctx.approved if r.get("stockId") != stock_id]
    removed = before - len(ctx.approved)
    save_json_atomic(ctx.paths.approved_out, ctx.approved)

    cleared = 0
    uploads = ctx.upload_state.setdefault("uploads", {})
    for rec in ctx.manifest.get(row, []):
        path = rec.get("local_path") or ""
        for label in PHOTO_LABELS:
            cache_key = hashlib.sha1(f"{path}:{label}".encode()).hexdigest()
            if cache_key in uploads:
                del uploads[cache_key]
                cleared += 1
    save_json_atomic(ctx.paths.upload_state, ctx.upload_state)
    entry = _queue_row(ctx, row)
    entry["resubmitCount"] = int(entry.get("resubmitCount") or 0) + 1
    if entry.get("submissionId"):
        entry["previousSubmissionId"] = entry["submissionId"]
    entry.pop("submissionId", None)
    entry.pop("adminReviewUrl", None)
    set_row_status(ctx, row, "pending", clear_issues=True)
    _append_history(
        entry,
        "reupload_reset",
        f"removed={removed} cache={cleared} resubmit=R{entry['resubmitCount']}",
    )
    save_queue(ctx)
    return {
        "ok": True,
        "row": row,
        "stockId": stock_id,
        "removedApproved": removed,
        "clearedUploadCache": cleared,
        "resubmitCount": entry["resubmitCount"],
        "nextSubmissionId": resolve_submission_id(row, entry),
        "message": (
            f"{stock_id} ready for re-upload — next submissionId "
            f"{resolve_submission_id(row, entry)} (photos + VIN powertrain + estimated FOB)"
        ),
    }


def process_row(
    ctx: PipelineContext,
    row: int,
    *,
    dry_run: bool = True,
    supplier_key: str | None = None,
    max_photos: int | None = None,
    ingest_knowledge: bool = True,
    upload_max_retries: int | None = None,
    upload_retry_backoff: float | None = None,
) -> dict[str, Any]:
    """Process one row: upload photos (if live), merge approved JSON, ingest knowledge."""
    prep = prepare_row(ctx, row, max_photos=max_photos)
    if not prep.get("ok"):
        return prep

    if dry_run:
        return prep

    key = supplier_key or os.environ.get("SUPPLIER_UPLOAD_KEY", "")
    if not key:
        return {
            "ok": False,
            "error": "SUPPLIER_UPLOAD_KEY required for live upload",
            "inspection": prep["inspection"],
        }

    fields = resolve_listing_fields(ctx, row)
    assert fields is not None
    photo_cap = max_photos if max_photos is not None else DEFAULT_UPLOAD_MAX_PHOTOS
    cap = _effective_max_photos(fields, photo_cap)
    token = _get_upload_token(
        ctx.paths.api_base,
        key,
        max_retries=upload_max_retries,
        retry_backoff=upload_retry_backoff,
    )
    uploaded: list[dict[str, Any]] = []
    uploads = ctx.upload_state.setdefault("uploads", {})

    for pick in fields["picks"][:cap]:
        path = Path(pick["path"])
        cache_key = hashlib.sha1(f"{path}:{pick['label']}".encode()).hexdigest()
        if cache_key not in uploads:
            uploads[cache_key] = _upload_file(
                ctx.paths.api_base,
                path,
                pick["label"],
                key,
                token,
                max_retries=upload_max_retries,
                retry_backoff=upload_retry_backoff,
            )
            save_json_atomic(ctx.paths.upload_state, ctx.upload_state)
        uploaded.append(uploads[cache_key])

    extras = _listing_build_extras(ctx, row, fields)
    record = build_record(
        fields["source"],
        fields["brand"],
        fields["brand_slug"],
        fields["model"],
        fields["model_key"],
        uploaded,
        fields["vin_info"],
        **extras,
    )
    _merge_approved(ctx, record)
    set_row_status(ctx, row, "uploaded")

    knowledge_result = None
    if ingest_knowledge:
        from tools import knowledge_ingest

        out_ref = str(ctx.paths.approved_out.relative_to(ctx.paths.root))
        knowledge_result = knowledge_ingest.ingest_listing_record(
            record,
            source="qxb_upload",
            source_ref=f"{out_ref}#{record['stockId']}",
        )
        set_row_status(ctx, row, "knowledge")

    return {
        "ok": True,
        "dryRun": False,
        "stockId": record["stockId"],
        "record": record,
        "knowledge": knowledge_result,
        "out": str(ctx.paths.approved_out),
        "audit": audit_row(ctx, row, record=record),
        "nextStep": f"/qxb submit-review {row}",
    }


def enrich_row_from_vin(ctx: PipelineContext, row: int) -> dict[str, Any]:
    """Re-decode VIN and patch local approved import JSON (engine/transmission/drivetrain)."""
    stock_id = f"QXB{row:04d}"
    rec = next((r for r in ctx.approved if r.get("stockId") == stock_id), None)
    if not rec:
        return {
            "ok": False,
            "error": f"{stock_id} not in {ctx.paths.approved_out} — run process --live first",
        }
    vin = str(rec.get("vin") or "").strip().upper()
    if len(vin) != 17:
        return {"ok": False, "error": f"invalid VIN on record: {vin or 'empty'}"}
    decode = decode_vin_via_api(ctx.paths.api_base, vin, root=ctx.paths.root)
    if not (
        decode.get("ok")
        and decode.get("engineCode")
        and decode.get("transmissionCode")
    ):
        fields = resolve_listing_fields(ctx, row)
        fb = fallback_vin_decode(ctx, fields) if fields else None
        if fb:
            decode = fb
        elif not decode.get("ok"):
            return {"ok": False, "error": decode.get("error"), "vin": vin, "stockId": stock_id}
        else:
            return {
                "ok": False,
                "error": "engineCode/transmissionCode missing after VIN decode",
                "vin": vin,
                "stockId": stock_id,
            }
    before = {
        "engineCode": rec.get("engineCode"),
        "transmissionCode": rec.get("transmissionCode"),
        "drivetrain": rec.get("drivetrain"),
    }
    apply_vin_decode_to_record(rec, decode)
    _merge_approved(ctx, rec)
    return {
        "ok": True,
        "stockId": stock_id,
        "vin": vin,
        "before": before,
        "after": {
            "engineCode": rec.get("engineCode"),
            "transmissionCode": rec.get("transmissionCode"),
            "drivetrain": rec.get("drivetrain"),
            "title": rec.get("title"),
        },
        "adminNote": (
            "若已 submit-review，请在 Admin Pending 里同步改发动机/变速箱字段后再批准"
        ),
    }


def format_enrich_report(data: dict[str, Any]) -> str:
    if not data.get("ok"):
        return f"VIN enrich 失败: {data.get('error', 'unknown')}"
    lines = [
        f"VIN enrich — {data.get('stockId')} ({data.get('vin')})",
        f"  engine: {data['before'].get('engineCode') or '—'} → {data['after'].get('engineCode')}",
        f"  trans:  {data['before'].get('transmissionCode') or '—'} → {data['after'].get('transmissionCode')}",
        f"  drive:  {data['before'].get('drivetrain') or '—'} → {data['after'].get('drivetrain')}",
        f"  title:  {data['after'].get('title')}",
    ]
    if data.get("adminNote"):
        lines.append(data["adminNote"])
    return "\n".join(lines)


def listing_to_submission(record: dict[str, Any]) -> dict[str, Any]:
    """Convert QXB import record → half-cut submission for admin review queue."""
    photos = []
    for p in record.get("photos") or []:
        if not isinstance(p, dict):
            continue
        url = str(p.get("url") or "").strip()
        if not url:
            continue
        label = PHOTO_LABEL_TO_REVIEW.get(p.get("label") or "", p.get("label") or "Photo")
        entry: dict[str, Any] = {"label": label, "url": url}
        if p.get("fileName"):
            entry["fileName"] = p["fileName"]
        if p.get("mimeType"):
            entry["mimeType"] = p["mimeType"]
        if p.get("size"):
            entry["size"] = p["size"]
        photos.append(entry)

    return {
        "submissionId": record.get("submissionId") or record.get("stockId"),
        "reviewStatus": "pending",
        "vin": normalize_vin_strict(record.get("vin") or ""),
        "brand": record.get("brand"),
        "brandSlug": record.get("brandSlug"),
        "model": record.get("model"),
        "year": record.get("year"),
        "mileage": record.get("mileage"),
        "engineCode": record.get("engineCode") or "",
        "transmissionCode": record.get("transmissionCode") or "",
        "drivetrain": record.get("drivetrain") or "2WD",
        "vehicleCondition": record.get("vehicleCondition") or "Half Cut",
        "vehicleCategory": record.get("vehicleCategory") or "passenger",
        "truckPartType": record.get("truckPartType") or "",
        "machineryType": record.get("machineryType") or "",
        "decodeMethod": record.get("decodeMethod") or "Manual Entry",
        "decodeConfidence": record.get("decodeConfidence"),
        "supplierName": record.get("supplierName") or "汽修宝",
        "supplierPhone": record.get("supplierPhone") or "",
        "supplierCity": record.get("supplierCity") or "",
        "supplierVerified": bool(record.get("supplierVerified")),
        "photos": photos,
        "video": None,
        "videoUrl": "",
        "inventoryStatus": record.get("status") or "Available",
        "priceUsd": record.get("priceUsd"),
        "priceEstimated": record.get("priceEstimated"),
        "notes": record.get("notes") or "",
        "includedParts": record.get("includedParts") or [],
        "shortDescription": record.get("shortDescription") or "",
        "qxbStockId": record.get("stockId"),
        "qxbSource": "apinventory_batch",
    }


def submit_row_for_review(
    ctx: PipelineContext,
    row: int,
    *,
    supplier_key: str | None = None,
    upload_max_retries: int | None = None,
    upload_retry_backoff: float | None = None,
) -> dict[str, Any]:
    """POST listing to production /api/half-cuts/submissions → admin pending review."""
    stock_id = f"QXB{row:04d}"
    rec = next((r for r in ctx.approved if r.get("stockId") == stock_id), None)
    if not rec:
        return {
            "ok": False,
            "error": f"{stock_id} not in {ctx.paths.approved_out} — run /qxb process {row} --live approved first",
        }

    key = supplier_key or os.environ.get("SUPPLIER_UPLOAD_KEY", "")
    if not key:
        return {"ok": False, "error": "SUPPLIER_UPLOAD_KEY required"}

    vin = normalize_vin_strict(rec.get("vin") or "")
    dup = find_duplicate_vin(ctx, vin, exclude_row=row) if vin else None
    if dup:
        msg = format_duplicate_vin_message(dup)
        return {"ok": False, "error": msg, "duplicate": dup, "stockId": stock_id}

    payload = listing_to_submission(rec)
    retry_opts = _upload_retry_opts(upload_max_retries, upload_retry_backoff)
    try:
        token = _get_upload_token(
            ctx.paths.api_base,
            key,
            max_retries=upload_max_retries,
            retry_backoff=upload_retry_backoff,
        )
        resp = _request_json(
            f"{ctx.paths.api_base}/api/half-cuts/submissions",
            "POST",
            {"X-Supplier-Key": key, "X-Upload-Token": token},
            payload,
            **retry_opts,
        )
    except RuntimeError as exc:
        msg = str(exc)
        if "Duplicate submission" in msg:
            set_row_status(ctx, row, "pending_review", issue="already on server review queue")
            return {
                "ok": True,
                "alreadySubmitted": True,
                "stockId": stock_id,
                "submissionId": payload.get("submissionId"),
                "adminReviewUrl": f"{ctx.paths.api_base}/admin/inventory.html?tab=pending",
                "message": "该 submission 已在审核队列中",
            }
        if "already exists" in msg.lower():
            return {
                "ok": False,
                "error": msg,
                "stockId": stock_id,
                "hint": "服务器已有同 VIN 的 pending 记录，请先在 Admin 处理旧单或改用 -R{n} 重提",
            }
        return {"ok": False, "error": msg, "stockId": stock_id}

    if not isinstance(resp, dict) or not resp.get("ok"):
        return {
            "ok": False,
            "error": f"submit API unexpected response: {resp!r}",
            "stockId": stock_id,
        }

    set_row_status(ctx, row, "pending_review")
    entry = _queue_row(ctx, row)
    entry["submissionId"] = resp.get("submissionId") or payload.get("submissionId")
    entry.pop("issues", None)
    entry["adminReviewUrl"] = f"{ctx.paths.api_base}/admin/inventory.html?tab=pending"
    save_queue(ctx)

    try:
        from tools import memory_tool

        memory_tool.remember(
            f"QXB {stock_id} submitted for CEO review → admin pending queue",
            category="inventory",
            source="apinventory",
            tags=["qxb", "submit-review"],
        )
    except Exception:
        pass

    return {
        "ok": True,
        "stockId": stock_id,
        "submissionId": entry.get("submissionId"),
        "adminReviewUrl": entry["adminReviewUrl"],
        "apiResponse": resp,
    }


def format_submit_review_report(data: dict[str, Any]) -> str:
    if not data.get("ok"):
        return f"提交审核失败: {data.get('error', 'unknown')}"
    lines = [
        f"已提交 CEO 审核队列 — {data.get('stockId')}",
        f"submissionId: {data.get('submissionId')}",
        f"管理后台: {data.get('adminReviewUrl')}",
    ]
    if data.get("alreadySubmitted"):
        lines.insert(0, "（该记录此前已提交过，仍在 pending 队列）")
    lines.append("请在 Admin → Inventory Hub → Pending 进行二次审核、批准上线。")
    return "\n".join(lines)


def _admin_fetch_state(api_base: str, admin_password: str) -> dict[str, Any]:
    """Login and fetch /api/half-cuts/state (read-only reconcile)."""
    import http.cookiejar

    jar = http.cookiejar.CookieJar()
    opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(jar))
    base = api_base.rstrip("/")
    login_req = urllib.request.Request(
        f"{base}/api/auth/login",
        data=json.dumps({"password": admin_password}).encode(),
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    with opener.open(login_req, timeout=30) as res:
        json.loads(res.read().decode())
    state_req = urllib.request.Request(f"{base}/api/half-cuts/state")
    with opener.open(state_req, timeout=30) as res:
        return json.loads(res.read().decode())


def _local_submission_rows(ctx: PipelineContext) -> list[dict[str, Any]]:
    """All QXB rows we believe were uploaded/submitted locally."""
    rows: dict[int, dict[str, Any]] = {}
    for rec in ctx.approved:
        stock_id = str(rec.get("stockId") or "")
        if not stock_id.startswith("QXB"):
            continue
        row_n = (rec.get("qxb") or {}).get("row")
        if row_n is None:
            try:
                row_n = int(stock_id[3:])
            except ValueError:
                continue
        q = (ctx.queue.get("rows") or {}).get(str(row_n), {})
        rows[row_n] = {
            "row": row_n,
            "stockId": stock_id,
            "submissionId": rec.get("submissionId") or q.get("submissionId"),
            "vin": rec.get("vin"),
            "brand": rec.get("brand"),
            "model": rec.get("model"),
            "localStatus": q.get("status"),
            "park": (q.get("park") or {}).get("category"),
        }
    for key, q in (ctx.queue.get("rows") or {}).items():
        row_n = int(key)
        if row_n in rows:
            continue
        if q.get("status") not in ("pending_review", "uploaded", "knowledge", "parked"):
            continue
        if q.get("submissionId") or str(q.get("status")) == "pending_review":
            rows[row_n] = {
                "row": row_n,
                "stockId": f"QXB{row_n:04d}",
                "submissionId": q.get("submissionId"),
                "vin": (ctx.vins.get(row_n) or {}).get("vin"),
                "brand": "",
                "model": "",
                "localStatus": q.get("status"),
                "park": (q.get("park") or {}).get("category"),
            }
    return sorted(rows.values(), key=lambda r: r["row"])


def reconcile_with_server(
    ctx: PipelineContext,
    *,
    admin_password: str | None = None,
) -> dict[str, Any]:
    """Compare local QXB submissions vs production Admin state."""
    local = _local_submission_rows(ctx)
    pwd = admin_password or os.environ.get("ADMIN_PASSWORD", "")
    if not pwd:
        return {
            "ok": False,
            "error": "ADMIN_PASSWORD required — set in .env then run /qxb reconcile",
            "localCount": len(local),
            "local": local,
        }
    try:
        state = _admin_fetch_state(ctx.paths.api_base, pwd)
    except Exception as exc:
        return {"ok": False, "error": str(exc), "localCount": len(local), "local": local}

    submissions = state.get("submissions") or []
    by_id = {
        str(s.get("submissionId")): s
        for s in submissions
        if s.get("submissionId")
    }
    local_ids = {str(r["submissionId"]) for r in local if r.get("submissionId")}

    on_server: list[dict[str, Any]] = []
    missing: list[dict[str, Any]] = []
    for loc in local:
        sid = str(loc.get("submissionId") or "")
        if sid and sid in by_id:
            srv = by_id[sid]
            on_server.append({
                **loc,
                "serverStatus": srv.get("reviewStatus"),
            })
        elif loc.get("localStatus") in ("pending_review", "uploaded", "knowledge", "parked"):
            missing.append(loc)

    server_only = [
        {
            "submissionId": s.get("submissionId"),
            "vin": s.get("vin"),
            "brand": s.get("brand"),
            "model": s.get("model"),
            "reviewStatus": s.get("reviewStatus"),
            "qxbStockId": s.get("qxbStockId"),
        }
        for s in submissions
        if str(s.get("submissionId") or "").startswith("QXB-")
        and str(s.get("submissionId")) not in local_ids
    ]

    return {
        "ok": True,
        "apiBase": ctx.paths.api_base,
        "localCount": len(local),
        "onServerCount": len(on_server),
        "missingCount": len(missing),
        "serverOnlyCount": len(server_only),
        "onServer": on_server,
        "missingOnServer": missing,
        "serverOnly": server_only,
        "reconciledAt": _iso_now(),
    }


def format_reconcile_report(data: dict[str, Any]) -> str:
    if not data.get("ok"):
        lines = [f"对账失败: {data.get('error')}"]
        if data.get("localCount"):
            lines.append(f"本地记录 {data['localCount']} 条 — 修好 ADMIN_PASSWORD 后重试")
        return "\n".join(lines)
    lines = [
        "QXB 本地 vs 服务器对账",
        f"API: {data.get('apiBase')}",
        f"本地已提交记录: {data.get('localCount')} | 服务器命中: {data.get('onServerCount')} | "
        f"漏提交: {data.get('missingCount')} | 仅服务器有: {data.get('serverOnlyCount')}",
        "",
    ]
    missing = data.get("missingOnServer") or []
    if missing:
        lines.append("【漏提交 — 需补推】")
        for m in missing:
            lines.append(
                f"  QXB{m['row']:04d} {m.get('submissionId') or '—'} "
                f"({m.get('brand')} {m.get('model')}) status={m.get('localStatus')} "
                f"park={m.get('park') or '—'}"
            )
        lines.append("")
    else:
        lines.append("✓ 无漏提交（本地 pending 均在服务器）")
        lines.append("")
    server_only = data.get("serverOnly") or []
    if server_only:
        lines.append("【仅服务器有 — 供核对】")
        for s in server_only[:20]:
            lines.append(f"  {s.get('submissionId')} {s.get('reviewStatus')} {s.get('brand')} {s.get('model')}")
        if len(server_only) > 20:
            lines.append(f"  … +{len(server_only) - 20} more")
    lines.append("")
    lines.append("补漏: /qxb unpark <row> → process --live → submit-review")
    return "\n".join(lines)


def _fetch_server_qxb_rows_via_ssh() -> dict[str, Any]:
    """Read production JSON over SSH when ADMIN_PASSWORD is not in local .env."""
    host = os.environ.get("QXB_PROD_SSH", "root@159.65.86.24")
    site = os.environ.get(
        "QXB_PROD_SITE_ROOT",
        "/root/.openclaw/workspace/inventory-site",
    )
    script = f"""import json, re
from pathlib import Path
root = Path({json.dumps(site)})
rows = set()

def add_stock(s):
    s = str(s or "").strip().upper()
    if s.startswith("QXB") and len(s) >= 7:
        try:
            rows.add(int(s[3:]))
        except ValueError:
            pass

def add_sid(s):
    m = re.match(r"^QXB-(\\d{{4}})", str(s or "").strip().upper())
    if m:
        rows.add(int(m.group(1)))

for name in ("half-cut-submissions.json", "half-cut-approved.json"):
    p = root / "data" / name
    if not p.exists():
        continue
    for item in json.load(p.open()):
        add_stock(item.get("qxbStockId") or item.get("stockId"))
        add_sid(item.get("submissionId"))
        qxb = item.get("qxb") or {{}}
        if qxb.get("row") is not None:
            try:
                rows.add(int(qxb["row"]))
            except (TypeError, ValueError):
                pass
print(json.dumps(sorted(rows)))
"""
    try:
        proc = subprocess.run(
            ["ssh", "-o", "BatchMode=yes", "-o", "ConnectTimeout=15", host, "python3", "-"],
            input=script,
            capture_output=True,
            text=True,
            timeout=45,
            check=False,
        )
        if proc.returncode != 0:
            err = (proc.stderr or proc.stdout or "ssh failed").strip()[:300]
            return {"ok": False, "error": err, "rows": set()}
        data = json.loads((proc.stdout or "[]").strip() or "[]")
        row_set = {int(x) for x in data}
        return {"ok": True, "rows": row_set, "count": len(row_set), "via": "ssh"}
    except Exception as exc:
        return {"ok": False, "error": str(exc), "rows": set()}


def fetch_server_qxb_rows(ctx: PipelineContext) -> dict[str, Any]:
    """QXB row numbers already on production (pending submissions or approved)."""
    pwd = os.environ.get("ADMIN_PASSWORD", "")
    if pwd:
        try:
            state = _admin_fetch_state(ctx.paths.api_base, pwd)
        except Exception as exc:
            return {"ok": False, "error": str(exc), "rows": set()}

        rows: set[int] = set()

        def _add_qxb_stock(stock_id: str) -> None:
            sid = str(stock_id or "").strip().upper()
            if sid.startswith("QXB") and len(sid) >= 7:
                try:
                    rows.add(int(sid[3:]))
                except ValueError:
                    pass

        def _add_submission_id(sub_id: str) -> None:
            text = str(sub_id or "").strip().upper()
            m = re.match(r"^QXB-(\d{4})", text)
            if m:
                rows.add(int(m.group(1)))

        for sub in state.get("submissions") or []:
            _add_qxb_stock(sub.get("qxbStockId") or sub.get("stockId"))
            _add_submission_id(sub.get("submissionId"))
        for item in state.get("approved") or []:
            _add_qxb_stock(item.get("qxbStockId") or item.get("stockId"))
            qxb = item.get("qxb") or {}
            if qxb.get("row") is not None:
                try:
                    rows.add(int(qxb["row"]))
                except (TypeError, ValueError):
                    pass

        return {"ok": True, "rows": rows, "count": len(rows), "fetchedAt": _iso_now(), "via": "admin"}

    ssh_result = _fetch_server_qxb_rows_via_ssh()
    if ssh_result.get("ok"):
        ssh_result["fetchedAt"] = _iso_now()
        return ssh_result
    return ssh_result


def sync_server_dedup(ctx: PipelineContext) -> dict[str, Any]:
    """Mark local queue rows that already exist on production; return skip set."""
    fetched = fetch_server_qxb_rows(ctx)
    if not fetched.get("ok"):
        return fetched

    server_rows: set[int] = fetched["rows"]
    synced = 0
    for row in server_rows:
        entry = (ctx.queue.get("rows") or {}).get(str(row))
        if not entry:
            continue
        if entry.get("status") not in ("pending", "inspected", "prepared", "uploaded", "knowledge"):
            continue
        set_row_status(ctx, row, "pending_review", issue="already on server (dedup sync)")
        synced += 1

    return {
        "ok": True,
        "serverRowCount": len(server_rows),
        "localSynced": synced,
        "rows": server_rows,
        "fetchedAt": fetched.get("fetchedAt"),
    }


def count_remaining_upload_rows(ctx: PipelineContext) -> dict[str, Any]:
    """Rows not yet in approved import — for bulk finish pass."""
    skip_status = {"uploaded", "knowledge", "pending_review", "blocked", "skipped", "parked"}
    server_rows: set[int] = set()
    if bulk_upload_mode():
        dedup = sync_server_dedup(ctx)
        if dedup.get("ok"):
            server_rows = dedup.get("rows") or set()
    min_manifest = 1 if bulk_upload_mode() else 3
    remaining: list[int] = []
    skipped_on_server: list[int] = []
    blocked_vin: list[int] = []
    for row in sorted(ctx.sources):
        if row in server_rows:
            skipped_on_server.append(row)
            continue
        entry = (ctx.queue.get("rows") or {}).get(str(row), {})
        if entry.get("status") in skip_status:
            if entry.get("status") == "parked" and (entry.get("park") or {}).get("category") == "no_vin":
                blocked_vin.append(row)
            continue
        if len(ctx.manifest.get(row, [])) < min_manifest:
            continue
        stock_id = f"QXB{row:04d}"
        if any(r.get("stockId") == stock_id for r in ctx.approved):
            continue
        remaining.append(row)
    return {
        "remaining": remaining,
        "remainingCount": len(remaining),
        "nextRow": remaining[0] if remaining else None,
        "parkedCount": sum(
            1 for e in (ctx.queue.get("rows") or {}).values() if e.get("status") == "parked"
        ),
        "noVinParked": blocked_vin,
        "skippedOnServer": skipped_on_server,
        "skippedOnServerCount": len(skipped_on_server),
        "bulkMode": bulk_upload_mode(),
    }


def format_remaining_report(data: dict[str, Any]) -> str:
    lines = [
        f"待上传行: {data.get('remainingCount')}（暂缓 {data.get('parkedCount')} 行本轮跳过）",
        f"下一行: {data.get('nextRow') or '—'}",
    ]
    rem = data.get("remaining") or []
    if rem:
        preview = ", ".join(str(r) for r in rem[:15])
        if len(rem) > 15:
            preview += f", … (+{len(rem) - 15})"
        lines.append(f"清单: {preview}")
    if data.get("noVinParked"):
        lines.append(f"无 VIN 暂缓: {data['noVinParked']}")
    lines.append("全部跑完后: /qxb reconcile 与服务器对账补漏")
    return "\n".join(lines)


def audit_row(
    ctx: PipelineContext,
    row: int,
    *,
    record: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Report where a row sits in the upload → catalog pipeline."""
    stock_id = f"QXB{row:04d}"
    rec = record or next((r for r in ctx.approved if r.get("stockId") == stock_id), None)
    result: dict[str, Any] = {
        "row": row,
        "stockId": stock_id,
        "importJson": str(ctx.paths.approved_out),
        "inImportJson": rec is not None,
        "liveCatalog": False,
        "photoCount": 0,
        "pendingPhotos": 0,
        "publicPhotos": 0,
        "samplePhotoUrl": None,
        "liveCatalogUrl": f"{ctx.paths.api_base}/half-cuts/",
        "adminUrl": f"{ctx.paths.api_base}/admin/inventory.html",
        "adminReviewUrl": f"{ctx.paths.api_base}/admin/inventory.html?tab=pending",
        "publicApi": f"{ctx.paths.api_base}/api/half-cuts/public",
    }
    queue_entry = (ctx.queue.get("rows") or {}).get(str(row)) or {}
    result["queueStatus"] = queue_entry.get("status")
    result["submissionId"] = queue_entry.get("submissionId") or (rec or {}).get("submissionId")
    if not rec:
        result["stage"] = "not_imported"
        result["message"] = "未找到本地 import 记录 — 尚未 process --live"
        return result

    photos = rec.get("photos") or []
    result["photoCount"] = len(photos)
    for p in photos:
        url = str(p.get("url") or "")
        if "/pending/" in url:
            result["pendingPhotos"] += 1
        elif "/uploads/" in url:
            result["publicPhotos"] += 1
        if not result["samplePhotoUrl"] and url:
            result["samplePhotoUrl"] = (
                url if url.startswith("http") else f"{ctx.paths.api_base}{url}"
            )

    try:
        req = urllib.request.Request(
            result["publicApi"],
            headers={"User-Agent": "AsiaPower-QXB-Audit/1.0"},
        )
        with urllib.request.urlopen(req, timeout=20) as res:
            payload = json.loads(res.read().decode())
        approved = payload.get("approved") if isinstance(payload, dict) else payload
        if isinstance(approved, list):
            slug = rec.get("slug")
            for item in approved:
                if item.get("stockId") == stock_id or (slug and item.get("slug") == slug):
                    result["liveCatalog"] = True
                    result["liveSlug"] = item.get("slug")
                    result["detailUrl"] = (
                        f"{ctx.paths.api_base}/half-cuts/detail.html?stock={stock_id}"
                    )
                    break
    except Exception as exc:
        result["liveCatalogError"] = str(exc)

    if result["liveCatalog"]:
        result["stage"] = "live"
        result["message"] = "已在官网 catalog 上线"
    elif result.get("queueStatus") == "pending_review" or queue_entry.get("submissionId"):
        result["stage"] = "pending_review"
        result["message"] = (
            "已提交管理后台 Pending 审核队列 — 请 CEO 在 Admin Inventory 批准"
        )
    elif result["inImportJson"] and result["pendingPhotos"] > 0:
        result["stage"] = "staged_pending_promote"
        result["message"] = (
            "照片已传 R2（pending），本地 import JSON 已有；"
            "尚未提交审核 — 运行 /qxb submit-review <row>"
        )
    elif result["inImportJson"]:
        result["stage"] = "import_only"
        result["message"] = "仅有本地 import JSON，未上线"
    return result


def format_audit_report(data: dict[str, Any]) -> str:
    lines = [
        f"QXB 上传审计 — {data.get('stockId')} (row {data.get('row')})",
        f"阶段: {data.get('stage')} — {data.get('message', '')}",
        "",
        "1. 本地 staging（子龙 process 写入）",
        f"   import JSON: {'✅' if data.get('inImportJson') else '❌'} {data.get('importJson')}",
        f"   照片: {data.get('photoCount')} 张 "
        f"(pending {data.get('pendingPhotos')}, public {data.get('publicPhotos')})",
    ]
    if data.get("samplePhotoUrl"):
        lines.append(f"   样例图: {data.get('samplePhotoUrl')[:120]}...")
    lines.extend([
        "",
        "2. CEO 审核队列（Pending）",
        f"   状态: {data.get('queueStatus') or '—'}",
        f"   submissionId: {data.get('submissionId') or '—'}",
        f"   审核页: {data.get('adminReviewUrl')}",
        "",
        "3. 官网公开 catalog（买家可见）",
        f"   已上线: {'✅' if data.get('liveCatalog') else '❌ 尚未上线'}",
        f"   目录页: {data.get('liveCatalogUrl')}",
    ])
    if data.get("detailUrl"):
        lines.append(f"   详情页: {data.get('detailUrl')}")
    lines.extend([
        "",
        "子龙完整链条: pick → preview → process --live → submit-review → CEO 批准 → 官网上线",
    ])
    if data.get("stage") == "staged_pending_promote":
        lines.append(f"下一步: /qxb submit-review {data.get('row')}")
    return "\n".join(lines)


PREVIEW_HTTP_PORT_BASE = 19876


def _slug_preview_name(label: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", (label or "photo").lower()).strip("-")
    return slug[:40] or "photo"


def _preview_port(row: int) -> int:
    return PREVIEW_HTTP_PORT_BASE + int(row)


def _is_port_open(port: int) -> bool:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
        sock.settimeout(0.2)
        return sock.connect_ex(("127.0.0.1", port)) == 0


def _start_preview_http_server(reports_dir: Path, port: int) -> bool:
    """Serve reports/ on localhost so preview images load reliably."""
    if _is_port_open(port):
        return True
    try:
        subprocess.Popen(
            [
                sys.executable,
                "-m",
                "http.server",
                str(port),
                "--bind",
                "127.0.0.1",
            ],
            cwd=str(reports_dir),
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            start_new_session=True,
        )
    except OSError:
        return False
    for _ in range(20):
        if _is_port_open(port):
            return True
        time.sleep(0.1)
    return False


def _open_preview_url(url: str) -> bool:
    if sys.platform == "darwin":
        try:
            subprocess.run(["open", url], check=False)
            return True
        except OSError:
            pass
    return False


def preview_row_photos(
    ctx: PipelineContext,
    row: int,
    *,
    open_browser: bool = True,
    include_all: bool = False,
) -> dict[str, Any]:
    """Build local HTML preview of upload photo slots (and optionally all manifest photos)."""
    fields = resolve_listing_fields(ctx, row)
    if not fields:
        return {"ok": False, "error": f"Excel row {row} not found"}

    stock_id = f"QXB{row:04d}"
    slots = fields["picks"]
    items: list[dict[str, str]] = [
        {"label": p["label"], "path": str(Path(p["path"]).resolve())}
        for p in slots
    ]
    if include_all:
        for rec in ctx.manifest.get(row, []):
            path = str(Path(rec["local_path"]).resolve())
            if not any(x["path"] == path for x in items):
                items.append({
                    "label": f"manifest #{rec.get('image_index', '?')}",
                    "path": path,
                })

    preview_dir = ctx.paths.root / "reports" / f"qxb-preview-row-{row:04d}"
    img_dir = preview_dir / "img"
    if img_dir.is_dir():
        shutil.rmtree(img_dir)
    img_dir.mkdir(parents=True, exist_ok=True)

    cards = []
    for idx, item in enumerate(items):
        path = item["path"]
        label = item["label"]
        if not Path(path).is_file():
            cards.append(
                f"<section class='card missing'><h3>{html.escape(label)}</h3>"
                f"<p>Missing: {html.escape(path)}</p></section>"
            )
            continue
        ext = Path(path).suffix.lower() or ".jpg"
        safe = f"{idx:02d}_{_slug_preview_name(label)}{ext}"
        dest = img_dir / safe
        shutil.copy2(path, dest)
        rel = f"img/{safe}"
        cards.append(
            f"<section class='card'><h3>{html.escape(label)}</h3>"
            f"<img src='{html.escape(rel)}' alt='{html.escape(label)}' loading='lazy'/>"
            f"<p class='path'>{html.escape(Path(path).name)}</p></section>"
        )

    page = f"""<!DOCTYPE html>
<html lang="zh"><head>
<meta charset="utf-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1"/>
<title>{stock_id} photo preview</title>
<style>
  body {{ font-family: system-ui, sans-serif; margin: 1rem; background: #111; color: #eee; }}
  h1 {{ font-size: 1.25rem; }}
  .meta {{ color: #aaa; margin-bottom: 1rem; }}
  .grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(280px, 1fr)); gap: 1rem; }}
  .card {{ background: #1e1e1e; border-radius: 8px; padding: 0.75rem; }}
  .card img {{ width: 100%; height: auto; border-radius: 4px; background: #000; }}
  .card.missing {{ border: 1px solid #a33; }}
  .path {{ font-size: 0.7rem; word-break: break-all; color: #888; }}
</style>
</head><body>
<h1>{stock_id} — {html.escape(fields['brand'])} {html.escape(fields['model'])}</h1>
<p class="meta">VIN: {html.escape((fields['vin_info'] or {}).get('vin') or '—')} |
Trim: {html.escape(fields['source']['trim'])} |
Upload slots: {len(slots)} | Generated {html.escape(_now())}</p>
<div class="grid">{''.join(cards)}</div>
</body></html>
"""
    out = preview_dir / "index.html"
    out.write_text(page, encoding="utf-8")

    # Legacy single-file path for bookmarks/scripts.
    legacy = ctx.paths.root / "reports" / f"qxb-preview-row-{row:04d}.html"
    legacy.write_text(
        f'<!DOCTYPE html><meta charset="utf-8">'
        f'<meta http-equiv="refresh" content="0;url=./qxb-preview-row-{row:04d}/index.html">'
        f'<p>Redirecting to <a href="./qxb-preview-row-{row:04d}/index.html">preview</a>… '
        f'If images are blank, open via local server (see /qxb preview help).</p>',
        encoding="utf-8",
    )

    port = _preview_port(row)
    reports_dir = ctx.paths.root / "reports"
    server_ok = _start_preview_http_server(reports_dir, port)
    preview_url = f"http://127.0.0.1:{port}/qxb-preview-row-{row:04d}/index.html"
    opened = bool(server_ok and open_browser and _open_preview_url(preview_url))

    return {
        "ok": True,
        "row": row,
        "stockId": stock_id,
        "html": str(out),
        "previewUrl": preview_url if server_ok else None,
        "slotCount": len(slots),
        "browserOpened": opened,
        "previewServer": server_ok,
        "previewPort": port if server_ok else None,
    }


def format_photo_preview_report(data: dict[str, Any]) -> str:
    if not data.get("ok"):
        return f"Photo preview failed: {data.get('error', 'unknown')}"
    lines = [
        f"Photo preview — {data.get('stockId')}",
        f"Folder: {Path(str(data.get('html') or '')).parent}",
        f"Upload slots: {data.get('slotCount')}",
    ]
    if data.get("previewUrl"):
        lines.append(f"浏览器打开: {data.get('previewUrl')}")
    if data.get("browserOpened"):
        lines.append("已在浏览器打开本地预览服务。")
    elif data.get("previewUrl"):
        lines.append("请复制上面的 http://127.0.0.1 链接到浏览器（不要双击 HTML 文件）。")
    else:
        lines.append("预览服务未启动 — 在终端运行:")
        port = data.get("previewPort") or PREVIEW_HTTP_PORT_BASE + int(data.get("row") or 0)
        lines.append(f"  cd reports && python3 -m http.server {port} --bind 127.0.0.1")
        row = int(data.get("row") or 0)
        lines.append(f"  然后打开 http://127.0.0.1:{port}/qxb-preview-row-{row:04d}/index.html")
    lines.append("Finder 原图文件夹: /qxb preview <row> open-folder")
    return "\n".join(lines)


def open_photo_folder(ctx: PipelineContext, row: int) -> str | None:
    photos = ctx.manifest.get(row, [])
    if not photos:
        return None
    folder = Path(photos[0]["local_path"]).resolve().parent
    if sys.platform == "darwin" and folder.is_dir():
        subprocess.run(["open", str(folder)], check=False)
        return str(folder)
    return str(folder) if folder.is_dir() else None


def queue_summary(ctx: PipelineContext) -> dict[str, Any]:
    """Counts by status and next actionable row."""
    rows = ctx.queue.get("rows") or {}
    counts: dict[str, int] = defaultdict(int)
    for entry in rows.values():
        counts[entry.get("status", "pending")] += 1

    total_excel = len(ctx.sources)
    with_photos = sum(1 for r in ctx.sources if len(ctx.manifest.get(r, [])) >= 3)
    approved_ids = {r.get("stockId") for r in ctx.approved}

    return {
        "xlsx": str(ctx.paths.xlsx),
        "xlsxExists": ctx.paths.xlsx.is_file(),
        "totalExcelRows": total_excel,
        "rowsWithPhotos": with_photos,
        "approvedRecords": len(ctx.approved),
        "queueCounts": dict(counts),
        "parkedCount": counts.get("parked", 0),
        "nextRow": find_next_row(ctx),
        "sources": {
            "manifest": str(ctx.paths.manifest),
            "approvedOut": str(ctx.paths.approved_out),
            "agentQueue": str(ctx.paths.agent_queue),
        },
    }


def find_next_row(ctx: PipelineContext) -> int | None:
    """First row that has photos and is not uploaded/blocked/skipped/parked/in review."""
    skip = {"uploaded", "knowledge", "blocked", "skipped", "parked", "pending_review"}
    for row in sorted(ctx.sources):
        entry = (ctx.queue.get("rows") or {}).get(str(row), {})
        if entry.get("status") in skip:
            continue
        if len(ctx.manifest.get(row, [])) < 3:
            continue
        stock_id = f"QXB{row:04d}"
        if stock_id in {r.get("stockId") for r in ctx.approved}:
            continue
        return row
    return None


def format_inspection_report(data: dict[str, Any]) -> str:
    lines = [
        f"QXB Row {data['row']} — {data.get('stockId') or 'N/A'}",
        f"Queue status: {data.get('status')}",
        f"Ready: {'yes' if data.get('ready') else 'no'}",
    ]
    if data.get("blockers"):
        lines.append("Blockers:")
        lines.extend(f"  - {b}" for b in data["blockers"])
    if data.get("warnings"):
        lines.append("Warnings:")
        lines.extend(f"  - {w}" for w in data["warnings"])
    fields = data.get("fields") or {}
    lines.extend([
        "",
        f"Brand: {fields.get('brand')} | Model: {fields.get('model')}",
        f"Trim: {fields.get('trim')}",
        f"VIN: {fields.get('vin') or '—'}",
        f"VIN OCR: {fields.get('vinConfidence') or '—'}",
        f"Engine: {fields.get('engineCode') or '—'} | Trans: {fields.get('transmissionCode') or '—'}",
        f"FOB est.: ${fields.get('priceUsd') or '—'}"
        + (" (estimated)" if fields.get('priceEstimated') else ""),
        f"Decode: {fields.get('decodeMethod') or '—'}",
        f"Photos: {fields.get('photoCount')} local → {len(fields.get('photoSlots') or [])} slots",
    ])
    for slot in fields.get("photoSlots") or []:
        lines.append(f"  - {slot['label']}: {slot['path']}")
    if data.get("issues"):
        lines.append("Recorded issues:")
        for issue in data["issues"]:
            lines.append(f"  - {issue.get('text', issue)}")
    return "\n".join(lines)


def format_status_report(summary: dict[str, Any]) -> str:
    lines = [
        "QXB Upload Queue (子龙 one-by-one workflow)",
        f"Pipeline version: {PIPELINE_VERSION}",
        f"Excel: {summary['xlsx']} ({'found' if summary['xlsxExists'] else 'MISSING'})",
        f"Rows in Excel: {summary['totalExcelRows']} | with ≥3 photos: {summary['rowsWithPhotos']}",
        f"Approved import records: {summary['approvedRecords']}",
    ]
    if summary.get("queueCounts"):
        lines.append("Queue status:")
        for status, count in sorted(summary["queueCounts"].items()):
            lines.append(f"  - {status}: {count}")
    nxt = summary.get("nextRow")
    lines.append(f"Next row to process: {nxt if nxt is not None else '(none — done or blocked)'}")
    parked = summary.get("parkedCount")
    if parked:
        lines.append(f"Parked (暂缓): {parked} — /qxb parked 查看清单")
    lines.append("")
    lines.append("Commands: /qxb next | inspect <row> | prepare <row> | process <row>")
    return "\n".join(lines)
